"""
Excel I/O operations for reading and writing Excel files.
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from .utils import get_color_fills, trim_string, is_empty, values_equal


def load_excel_file(file_path):
    """Load Excel file and return dictionary of dataframes by sheet name."""
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if not file_path.lower().endswith('.xlsx'):
            raise ValueError(f"File must be .xlsx format: {file_path}")
        
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        sheets_data = {}
        
        for sheet_name in excel_file.sheet_names:
            try:
                # Always use header=2 for consistency (based on demo files structure)
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=2)
                df.columns = [str(col).strip() for col in df.columns]
                
                # Clean up the dataframe
                df = df.dropna(how='all')  # Remove completely empty rows
                sheets_data[sheet_name] = df
                
            except Exception as e:
                print(f"Warning: Could not load sheet '{sheet_name}' from {file_path}: {e}")
                sheets_data[sheet_name] = pd.DataFrame()
        
        return sheets_data
        
    except Exception as e:
        raise Exception(f"Error loading Excel file {file_path}: {e}")


def save_comparison_excel(file_path, comparison_data, file_a_name, file_b_name):
    """Save comparison results to Excel with color coding."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    colors = get_color_fills()
    
    for sheet_name, sheet_data in comparison_data.items():
        if sheet_data.empty:
            continue
            
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(sheet_data.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data with color coding
        for row_idx, (_, row) in enumerate(sheet_data.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply color based on status
                if col_name == "Status":
                    if value == "A-ONLY":
                        cell.fill = colors['A_ONLY']
                    elif value == "B-ONLY":
                        cell.fill = colors['B_ONLY']
                    elif value == "CHANGED":
                        cell.fill = colors['CHANGED']
                elif "(File A)" in col_name or "(File B)" in col_name:
                    # Check if this is a changed value
                    base_col = col_name.replace(" (File A)", "").replace(" (File B)", "")
                    if base_col in sheet_data.columns:
                        # Find the corresponding A and B columns
                        col_a = f"{base_col} (File A)"
                        col_b = f"{base_col} (File B)"
                        
                        if col_a in sheet_data.columns and col_b in sheet_data.columns:
                            row_a_val = row.get(col_a)
                            row_b_val = row.get(col_b)
                            
                            if not values_equal(row_a_val, row_b_val) and row.get("Status") == "CHANGED":
                                cell.fill = colors['CHANGED']
                            elif col_name == col_a and row.get("Status") == "A-ONLY":
                                cell.fill = colors['A_ONLY']
                            elif col_name == col_b and row.get("Status") == "B-ONLY":
                                cell.fill = colors['B_ONLY']
    
    wb.save(file_path)


def get_sheet_headers(file_path, sheet_name):
    """Get headers for a specific sheet from a file."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=2)
        return list(df.columns)
    except:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
            return list(df.columns)
        except:
            return []
