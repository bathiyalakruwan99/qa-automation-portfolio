"""
Utility functions for Excel comparison tool.
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill


def ensure_directories():
    """Ensure output directories exist."""
    os.makedirs('reports', exist_ok=True)
    os.makedirs('outputs', exist_ok=True)


def get_timestamp():
    """Get current timestamp in YYYYMMDD_HHMMSS format."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def trim_string(value):
    """Trim whitespace from string values."""
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def is_empty(value):
    """Check if value is empty (None, NaN, or empty string)."""
    if pd.isna(value) or value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def values_equal(val1, val2):
    """Compare two values for equality, handling empty values and trimming."""
    val1_trimmed = trim_string(val1)
    val2_trimmed = trim_string(val2)
    
    # Both empty
    if is_empty(val1_trimmed) and is_empty(val2_trimmed):
        return True
    
    # One empty, one not
    if is_empty(val1_trimmed) or is_empty(val2_trimmed):
        return False
    
    # Both non-empty, compare exactly
    return val1_trimmed == val2_trimmed


def get_color_fills():
    """Get color fill definitions for Excel export."""
    return {
        'A_ONLY': PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid"),  # Blue
        'B_ONLY': PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),  # Green
        'CHANGED': PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Red
    }


def create_composite_key(row, key_columns):
    """Create composite key from row data and key columns."""
    key_parts = []
    for col in key_columns:
        if col in row.index:
            value = trim_string(row[col])
            if is_empty(value):
                return None  # Invalid key if any key field is empty
            key_parts.append(str(value))
        else:
            return None  # Invalid key if column doesn't exist
    return tuple(key_parts)


def find_duplicates(df, key_columns):
    """Find duplicate keys in a dataframe."""
    if df.empty:
        return set()
    
    # Create composite keys
    keys = []
    for _, row in df.iterrows():
        key = create_composite_key(row, key_columns)
        if key is not None:
            keys.append(key)
    
    # Find duplicates
    seen = set()
    duplicates = set()
    for key in keys:
        if key in seen:
            duplicates.add(key)
        else:
            seen.add(key)
    
    return duplicates


def validate_headers(df, expected_headers, sheet_name, file_name):
    """Validate that headers match expected order."""
    actual_headers = list(df.columns)
    
    if actual_headers == expected_headers:
        return True, "OK"
    else:
        return False, f"HEADER_ORDER_MISMATCH\nExpected: {expected_headers}\nActual: {actual_headers}"


# Sheet configuration
SHEET_CONFIG = {
    "1 - Organization Details": {
        "key_columns": ["Organization Short Name"],
        "display_name": "Organization Details"
    },
    "2 - Divisions": {
        "key_columns": ["Organization Short Name"],
        "display_name": "Divisions"
    },
    "3 - Human Resources": {
        "key_columns": ["Organization Short Name", "NIC"],
        "display_name": "Human Resources"
    },
    "4 - Vehicles": {
        "key_columns": ["Organization Short Name", "Vehicle Number"],
        "display_name": "Vehicles"
    },
    "5 - Locations": {
        "key_columns": ["Organization Short Name", "Location Reference ID"],
        "display_name": "Locations"
    }
}
