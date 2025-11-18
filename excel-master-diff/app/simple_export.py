"""
Simple export functionality to create files in the style of the expected output.
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font


def export_simple_format(files_data, file_a_name, file_b_name, output_dir="outputs"):
    """Export files in the simple format matching the expected output style."""
    os.makedirs(output_dir, exist_ok=True)
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"simple_export_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    file_names = list(files_data.keys())
    file_a, file_b = file_names[0], file_names[1]
    
    # Get all unique sheet names from both files
    all_sheets = set()
    for file_sheets in files_data.values():
        all_sheets.update(file_sheets.keys())
    
    for sheet_name in sorted(all_sheets):
        print(f"Processing sheet: {sheet_name}")
        
        # Get data from both files
        df_a = files_data[file_a].get(sheet_name, pd.DataFrame())
        df_b = files_data[file_b].get(sheet_name, pd.DataFrame())
        
        if df_a.empty and df_b.empty:
            continue
        
        # Create worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Get all unique columns from both files
        all_columns = set()
        if not df_a.empty:
            all_columns.update(df_a.columns)
        if not df_b.empty:
            all_columns.update(df_b.columns)
        
        # Create simple format dataframe
        simple_rows = []
        
        # Get the maximum number of rows between both files
        max_rows = max(len(df_a), len(df_b)) if not df_a.empty and not df_b.empty else max(len(df_a), len(df_b))
        
        for i in range(max_rows):
            row_data = {}
            
            # Add all columns with 1st file and 2nd file suffixes
            for col in sorted(all_columns):
                # Get values from both files (use None if index out of range)
                val_a = df_a.iloc[i][col] if not df_a.empty and i < len(df_a) and col in df_a.columns else None
                val_b = df_b.iloc[i][col] if not df_b.empty and i < len(df_b) and col in df_b.columns else None
                
                row_data[f"{col} 1st file"] = val_a
                row_data[f"{col} 2nd file"] = val_b
            
            # Add status column (simple status)
            row_data["Status"] = "NON_BOI"  # Default status as in the example
            
            simple_rows.append(row_data)
        
        # Create simple dataframe
        simple_df = pd.DataFrame(simple_rows)
        
        if simple_df.empty:
            continue
        
        # Write headers
        for col_idx, col_name in enumerate(simple_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        
        # Write data
        for row_idx, (_, row) in enumerate(simple_df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  ✅ Created {len(simple_df)} rows with {len(simple_df.columns)} columns")
    
    wb.save(filepath)
    return filepath
