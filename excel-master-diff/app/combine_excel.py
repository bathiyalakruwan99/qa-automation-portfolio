"""
Excel export functionality for combining data from both files into comprehensive sheets.
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font


def export_combined_excel(files_data, file_a_name, file_b_name, output_dir="outputs"):
    """Export combined data from both files into comprehensive sheets."""
    os.makedirs(output_dir, exist_ok=True)
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"combined_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    file_names = list(files_data.keys())
    file_a, file_b = file_names[0], file_names[1]
    
    # Get all unique sheet names from both files
    all_sheets = set()
    for file_sheets in files_data.values():
        all_sheets.update(file_sheets.keys())
    
    for sheet_name in sorted(all_sheets):
        print(f"Processing sheet: {sheet_name}")
        
        # Get data from both files
        df_a = files_data[file_a].get(sheet_name, pd.DataFrame())
        df_b = files_data[file_b].get(sheet_name, pd.DataFrame())
        
        if df_a.empty and df_b.empty:
            continue
        
        # Create combined worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Get all unique columns from both files
        all_columns = set()
        if not df_a.empty:
            all_columns.update(df_a.columns)
        if not df_b.empty:
            all_columns.update(df_b.columns)
        
        # Create combined dataframe
        combined_rows = []
        
        # Add data from file A
        if not df_a.empty:
            for _, row in df_a.iterrows():
                row_data = {}
                for col in sorted(all_columns):
                    row_data[f"{col} (File A)"] = row.get(col) if col in row.index else None
                row_data["Source File"] = "File A"
                combined_rows.append(row_data)
        
        # Add data from file B
        if not df_b.empty:
            for _, row in df_b.iterrows():
                row_data = {}
                for col in sorted(all_columns):
                    row_data[f"{col} (File B)"] = row.get(col) if col in row.index else None
                row_data["Source File"] = "File B"
                combined_rows.append(row_data)
        
        # Create combined dataframe
        combined_df = pd.DataFrame(combined_rows)
        
        if combined_df.empty:
            continue
        
        # Write headers
        for col_idx, col_name in enumerate(combined_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        
        # Write data
        for row_idx, (_, row) in enumerate(combined_df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  ✅ Combined {len(combined_df)} rows with {len(combined_df.columns)} columns")
    
    wb.save(filepath)
    return filepath
