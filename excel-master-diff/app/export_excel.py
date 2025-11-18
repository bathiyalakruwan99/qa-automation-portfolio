"""
Excel export functionality for two-file comparisons with full column duplication.
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font


def export_comparison_excel(comparison_data, file_a_name, file_b_name, output_dir="outputs"):
    """Export comparison results to Excel with full column duplication (no color coding)."""
    os.makedirs(output_dir, exist_ok=True)
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"compare_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for sheet_name, sheet_df in comparison_data.items():
        if sheet_df.empty:
            continue
        
        # Create worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(sheet_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        
        # Write data without color coding
        for row_idx, (_, row) in enumerate(sheet_df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # No color coding applied - just show all data side by side
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filepath)
    return filepath
