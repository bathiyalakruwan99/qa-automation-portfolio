import argparse
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

SHEET_KEYS = {
    "1 - Organization Details": ["Organization Short Name"],
    "2 - Divisions": ["Organization Short Name"],
    "3 - Human Resources": ["Organization Short Name", "NIC"],
    "4 - Vehicles": ["Organization Short Name", "Vehicle Number"],
    "5 - Locations": ["Organization Short Name", "Location Reference ID"],
}

BLUE  = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")  # A-only (left)
GREEN = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")  # B-only (right)
RED   = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Changed (both)

def detect_header_row(xlsx_path: Path, sheet_name: str, expected_keys: list[str]) -> int:
    """Detect the header row using heuristics."""
    try:
        peek = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, nrows=15, engine="openpyxl")
    except Exception:
        return 0
    
    # 1) First row within top 15 that contains all key columns
    for ridx in range(min(15, len(peek))):
        vals = [str(x).strip() for x in list(peek.iloc[ridx].values)]
        if all(k in vals for k in expected_keys):
            return ridx
    
    # 2) Row with >= half of the key columns
    for ridx in range(min(15, len(peek))):
        vals = [str(x).strip() for x in list(peek.iloc[ridx].values)]
        if sum(1 for k in expected_keys if k in vals) >= max(1, len(expected_keys)//2):
            return ridx
    
    # 3) Row among top 15 with the most non-empty cells
    non_null = peek.notna().sum(axis=1)
    return int(non_null.idxmax()) if len(non_null) else 0

def read_table(xlsx_path: Path, sheet_name: str, expected_keys: list[str]) -> pd.DataFrame | None:
    """Read Excel sheet with proper header detection."""
    hdr = detect_header_row(xlsx_path, sheet_name, expected_keys)
    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=hdr, engine="openpyxl")
    except Exception:
        return None
    
    if df is not None:
        # Remove placeholder "Unnamed" columns
        df = df.loc[:, ~df.columns.to_series().astype(str).str.startswith("Unnamed")]
    
    return df

def normalize_strings(df: pd.DataFrame | None) -> pd.DataFrame | None:
    """Normalize string values for comparison while preserving original in output."""
    if df is None:
        return None
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_string_dtype(df[col]):
            s = df[col].astype(str).str.strip()
            df[col] = s.replace({"nan": np.nan})
    return df

def is_empty(v) -> bool:
    """Check if value is empty (None, NaN, or empty string)."""
    if v is None: 
        return True
    if isinstance(v, float) and np.isnan(v): 
        return True
    if isinstance(v, str) and v.strip() == "": 
        return True
    return False

def norm(v):
    """Normalize value for comparison."""
    return v.strip() if isinstance(v, str) else v

def build_map_and_order(df: pd.DataFrame | None, keys: list[str]):
    """Build mapping from composite key to rows and preserve original order."""
    order = []
    rows_map: dict[tuple, list[dict]] = {}
    
    if df is None or df.empty:
        return rows_map, order
    
    for _, row in df.iterrows():
        key = tuple(row.get(k, np.nan) for k in keys)
        rows_map.setdefault(key, []).append(row.to_dict())
        order.append(key)
    
    return rows_map, order

def write_sheet(ws, header_labels, rows, fills):
    """Write data to worksheet with formatting."""
    # Write headers
    ws.append(header_labels)
    ws.freeze_panes = "A2"
    
    # Write data with color fills
    for out_row, fill_plan in zip(rows, fills):
        ws.append(out_row)
        r = ws.max_row
        for pair_idx, (lf, rf) in enumerate(fill_plan):
            c_left  = 1 + pair_idx*2
            c_right = c_left + 1
            if lf is not None: 
                ws.cell(row=r, column=c_left ).fill = lf
            if rf is not None: 
                ws.cell(row=r, column=c_right).fill = rf
    
    # AutoFilter + column widths
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col and max_row:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    for ci in range(1, max_col+1):
        hv = ws.cell(row=1, column=ci).value
        width = min(max(12, len(str(hv)) + 2), 75)
        ws.column_dimensions[get_column_letter(ci)].width = width

def compare_files(first: Path, second: Path, out_path: Path):
    """Main comparison function."""
    a_label = first.name
    b_label = second.name

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, key_cols in SHEET_KEYS.items():
        print(f"Processing sheet: {sheet_name}")
        
        # Read tables (B first to preserve its order)
        df_b = read_table(second, sheet_name, key_cols)  # preserve B order
        df_a = read_table(first,  sheet_name, key_cols)
        
        if df_a is None and df_b is None:
            print(f"  [WARN] Sheet '{sheet_name}' not found in either file")
            continue

        # Normalize for comparison
        df_a = normalize_strings(df_a) if df_a is not None else pd.DataFrame(columns=key_cols)
        df_b = normalize_strings(df_b) if df_b is not None else pd.DataFrame(columns=key_cols)

        # Ensure key columns exist
        for k in key_cols:
            if k not in df_a.columns: 
                df_a[k] = np.nan
            if k not in df_b.columns: 
                df_b[k] = np.nan

        # Column order: B first, then A-only columns
        cols_b = list(df_b.columns)
        cols_a = list(df_a.columns)
        all_cols = cols_b + [c for c in cols_a if c not in cols_b]

        # Build maps + original row orders
        map_b, order_b = build_map_and_order(df_b, key_cols)
        map_a, order_a = build_map_and_order(df_a, key_cols)

        # Final row order: all keys as they appear in B, then A-only keys in A order
        seen = set()
        final_key_order = []
        for k in order_b:
            if k not in seen:
                final_key_order.append(k)
                seen.add(k)
        for k in order_a:
            if k not in seen:
                final_key_order.append(k)
                seen.add(k)

        # Assemble rows + fill plans
        out_rows = []
        fill_plans = []
        
        for key in final_key_order:
            rows_b = map_b.get(key, [])
            rows_a = map_a.get(key, [])
            max_len = max(len(rows_b), len(rows_a)) if max(len(rows_b), len(rows_a)) > 0 else 1

            for i in range(max_len):
                rb = rows_b[i] if i < len(rows_b) else None
                ra = rows_a[i] if i < len(rows_a) else None

                out_row = []
                fill_plan = []
                
                for col in all_cols:
                    va = ra.get(col, np.nan) if ra is not None else np.nan
                    vb = rb.get(col, np.nan) if rb is not None else np.nan

                    # Decide fills based on comparison rules
                    if ra is None and rb is not None:
                        lf, rf = (None, GREEN)       # B-only row
                    elif rb is None and ra is not None:
                        lf, rf = (BLUE, None)        # A-only row
                    elif ra is not None and rb is not None:
                        ea, eb = is_empty(va), is_empty(vb)
                        if ea and eb:
                            lf, rf = (None, None)    # Both empty
                        else:
                            if norm(va) == norm(vb):
                                lf, rf = (None, None)  # Same
                            else:
                                lf, rf = (RED, RED)    # Changed
                    else:
                        lf, rf = (None, None)

                    out_row.extend([va, vb])
                    fill_plan.append((lf, rf))

                out_rows.append(out_row)
                fill_plans.append(fill_plan)

        # Write sheet
        ws = wb.create_sheet(title=sheet_name[:31])
        header_labels = []
        for col in all_cols:
            header_labels.append(f"{col} ({a_label})")
            header_labels.append(f"{col} ({b_label})")
        
        write_sheet(ws, header_labels, out_rows, fill_plans)
        print(f"  [OK] Created sheet with {len(out_rows)} rows and {len(all_cols)} column pairs")

    # Save workbook
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[OK] Saved comparison to: {out_path}")

def main():
    """CLI entry point."""
    p = argparse.ArgumentParser(description="Excel flat side-by-side comparator")
    p.add_argument("--first", required=True, help="Path to 1st (left) Excel file")
    p.add_argument("--second", required=True, help="Path to 2nd (right) Excel file (row+column order baseline)")
    p.add_argument("--out", required=True, help="Path to output .xlsx")
    args = p.parse_args()
    
    compare_files(Path(args.first), Path(args.second), Path(args.out))

if __name__ == "__main__":
    main()
