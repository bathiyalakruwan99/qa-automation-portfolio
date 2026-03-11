"""
Export ticket history from ticket_data JSON files to Excel.
Reads all ticket JSON files, extracts key fields and status history, writes to xlsx.
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TICKETS_DIR = os.path.join(_PROJECT_ROOT, 'ticket_history', 'ticket_data')
SAMPLES_DIR = os.path.join(_PROJECT_ROOT, 'samples')


def get_status_name(ticket: dict) -> str:
    """Get current status as string."""
    status = ticket.get('status')
    if status is None:
        return ''
    if isinstance(status, dict):
        return (status.get('name') or '').strip()
    return str(status).strip()


def get_type_name(ticket: dict) -> str:
    """Get issue type as string."""
    it = ticket.get('issue_type') or ticket.get('issuetype')
    if it is None:
        return ''
    if isinstance(it, dict):
        return (it.get('name') or '').strip()
    return str(it).strip()


def get_priority_name(ticket: dict) -> str:
    """Get priority as string."""
    p = ticket.get('priority')
    if p is None:
        return ''
    if isinstance(p, dict):
        return (p.get('name') or '').strip()
    return str(p).strip()


def format_date(ts: str) -> str:
    """Format ISO timestamp to YYYY-MM-DD."""
    if not ts:
        return ''
    try:
        dt = datetime.fromisoformat(ts.replace('Z', '+00:00')[:19])
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return ts[:10] if len(ts) >= 10 else ts


def get_changelogs(ticket: dict) -> list:
    """Get changelog entries (status changes)."""
    logs = ticket.get('changelogs', [])
    if logs:
        return logs
    return ticket.get('changelog', {}).get('histories', [])


def find_status_transitions(ticket: dict) -> list:
    """Extract status transitions from changelog."""
    transitions = []
    for changelog in sorted(get_changelogs(ticket), key=lambda x: x.get('created', '')):
        created = changelog.get('created', '')
        for item in changelog.get('items', []):
            if item.get('field') == 'status':
                from_val = item.get('from_string') or item.get('fromString') or ''
                to_val = item.get('to_string') or item.get('toString') or ''
                if from_val or to_val:
                    transitions.append((format_date(created), from_val, to_val))
    return transitions


def load_all_tickets() -> list:
    """Load all ticket JSON files."""
    tickets = []
    if not os.path.exists(TICKETS_DIR):
        return tickets
    for fname in sorted(os.listdir(TICKETS_DIR)):
        if not fname.endswith('.json'):
            continue
        path = os.path.join(TICKETS_DIR, fname)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                tickets.append(json.load(f))
        except Exception as e:
            print(f"Warning: skip {fname}: {e}")
    return tickets


def main():
    os.makedirs(SAMPLES_DIR, exist_ok=True)
    tickets = load_all_tickets()
    if not tickets:
        print(f"No tickets in {TICKETS_DIR}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ticket History'

    # Headers
    headers = [
        'Ticket ID', 'Summary', 'Type', 'Priority', 'Status', 'Created', 'Updated',
        'Status Change Date', 'From Status', 'To Status'
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    row = 2
    for ticket in tickets:
        key = ticket.get('key', '')
        summary = (ticket.get('summary') or '').strip()
        issue_type = get_type_name(ticket)
        priority = get_priority_name(ticket)
        status = get_status_name(ticket)
        created = format_date(ticket.get('created', ''))
        updated = format_date(ticket.get('updated', ''))

        transitions = find_status_transitions(ticket)
        if transitions:
            for change_date, from_s, to_s in transitions:
                ws.cell(row, 1, key)
                ws.cell(row, 2, summary)
                ws.cell(row, 3, issue_type)
                ws.cell(row, 4, priority)
                ws.cell(row, 5, status)
                ws.cell(row, 6, created)
                ws.cell(row, 7, updated)
                ws.cell(row, 8, change_date)
                ws.cell(row, 9, from_s)
                ws.cell(row, 10, to_s)
                row += 1
        else:
            ws.cell(row, 1, key)
            ws.cell(row, 2, summary)
            ws.cell(row, 3, issue_type)
            ws.cell(row, 4, priority)
            ws.cell(row, 5, status)
            ws.cell(row, 6, created)
            ws.cell(row, 7, updated)
            ws.cell(row, 8, '')
            ws.cell(row, 9, '')
            ws.cell(row, 10, '')
            row += 1

    out_path = os.path.join(SAMPLES_DIR, 'ticket_history_sample.xlsx')
    wb.save(out_path)
    print(f"Written {out_path} with {row - 2} rows")


if __name__ == '__main__':
    main()
