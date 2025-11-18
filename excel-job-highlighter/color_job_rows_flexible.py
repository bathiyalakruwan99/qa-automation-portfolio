import openpyxl
from openpyxl.styles import PatternFill
import re

def color_job_rows_flexible():
    """
    Color code rows in Excel files with flexible matching for job IDs
    """
    
    # Define colors
    GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    
    # Green patterns - Jobs found in BOTH files
    green_patterns = [
        # Rhino patterns
        r'Rhino.*July.*2025', r'Rhino.*Aug.*6', r'Rhino.*Aug.*7', r'Rhino.*Aug.*11', r'Rhino.*Aug.*13', r'Rhino.*Aug.*14',
        r'203629.*Rhino.*Aug.*6', r'203610.*Rhino.*Aug.*7', r'202432.*Rhino.*Aug.*7', r'203627.*Rhino.*Aug.*7',
        r'203639.*Rhino.*Aug.*7', r'203685.*Rhino.*Aug.*7', r'Rhino.*202440.*Aug.*6',
        
        # LANWA patterns
        r'LAN\d+.*Lanwa.*Aug', r'Lanwa.*Aug.*5', r'Lanwa.*Aug.*11', r'Lanwa.*Aug.*12', r'Lanwa.*Aug.*14',
        r'TO\d+.*Lanwa.*Aug',
        
        # Sumithra patterns
        r'SH\d+.*July', r'SH\d+.*August.*2025', r'SH\d+.*New.*August.*2025',
        
        # Sumithra Garments patterns
        r'SP.*July.*2025', r'SP\d+.*July.*2025', r'SP\d+.*August.*2025', r'SP\d+.*New.*August.*2025',
        r'SW\d+.*July.*2025', r'SW\d+.*August.*2025', r'SWA\d+.*July.*2025', r'SWA\d+.*August.*2025',
        r'SWA\d+.*Balance.*Commitment.*KM.*July.*2025',
        
        # Causeway patterns
        r'1010\d{7}', r'1010\d{6}',
        
        # INSEE patterns
        r'Insee.*July.*2025', r'Insee.*August.*2025',
        
        # Cisco patterns
        r'Cisco', r'Kithsiri.*LM-6923.*Cisco.*trips.*CS\d+', r'LB-1882.*Cisco.*\d+\.\d+\.\d+',
        r'LL-2883.*Cisco.*\d+\.\d+\.\d+', r'Shelton.*LI-2901.*Cisco.*trips',
        
        # Little Lion patterns
        r'TR\d+.*\d+/\d+', r'WH\d+.*\d+/\d+',
        
        # Kenilworth patterns
        r'Pasindu.*Kenilworth.*KWD\d+',
        
        # Siam City patterns
        r'DIMO.*July.*25.*2025', r'INSEE.*July.*25.*2025', r'Insee.*July.*25.*2025', r'Insee.*Agust.*4.*2025',
        r'SA00\d+'
    ]
    
    # Light blue patterns - Jobs found ONLY in jobs.md
    light_blue_patterns = [
        # Vijith Cisco patterns
        r'Vijith.*LB-1882.*Cisco.*trips', r'Vijith.*LL-2883.*Cisco.*trips', r'Vijith.*LL-2883.*Cisco.*\d+\.\d+\.\d+',
        
        # Shelton patterns
        r'Shelton.*LI-2901.*Cisco.*trips.*CS\d+', r'Shelton.*LI-2901.*Cisco.*\d+\.\d+\.\d+',
        
        # Additional Little Lion patterns
        r'TR2507.*\d+/\d+', r'TR2508.*\d+/\d+',
        
        # Additional patterns
        r'Causeway_28\.07_Payments', r'PQ 2456', r'SH-July-2025', r'Eco July 30 2025',
        r'101028\d+', r'101029\d+', r'101030\d+'
    ]
    
    try:
        # Load the workbook
        print("Loading sample_jobs.xlsx...")
        workbook = openpyxl.load_workbook('sample_jobs.xlsx')
        sheet = workbook.active
        print(f"Working with sheet: {sheet.title}")
        
        # Find job ID column
        job_id_column = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and any(keyword in str(cell_value).lower() for keyword in ['job', 'id', 'code', 'reference', 'order']):
                job_id_column = col
                print(f"Found job ID column: {col} - '{cell_value}'")
                break
        
        if job_id_column is None:
            print("Could not identify job ID column. Please check the file structure.")
            return
        
        # Process rows
        green_count = 0
        light_blue_count = 0
        total_rows = 0
        
        print(f"Processing rows with flexible pattern matching...")
        
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=job_id_column).value
            if cell_value:
                total_rows += 1
                cell_value_str = str(cell_value).strip()
                
                # Check patterns
                is_green = any(re.search(pattern, cell_value_str, re.IGNORECASE) for pattern in green_patterns)
                is_light_blue = any(re.search(pattern, cell_value_str, re.IGNORECASE) for pattern in light_blue_patterns)
                
                if is_green:
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = GREEN_FILL
                    green_count += 1
                    print(f"Row {row}: GREEN - '{cell_value_str}'")
                    
                elif is_light_blue:
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = LIGHT_BLUE_FILL
                    light_blue_count += 1
                    print(f"Row {row}: LIGHT BLUE - '{cell_value_str}'")
        
        # Save
        output_filename = 'jobs_colored_flexible.xlsx'
        workbook.save(output_filename)
        print(f"\nColoring completed!")
        print(f"Total rows processed: {total_rows}")
        print(f"Green rows: {green_count}")
        print(f"Light blue rows: {light_blue_count}")
        print(f"Output saved as: {output_filename}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    color_job_rows_flexible()
