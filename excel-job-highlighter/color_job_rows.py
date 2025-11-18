import openpyxl
from openpyxl.styles import PatternFill
import re

def color_job_rows():
    """
    Color code rows in Excel files based on job ID categories:
    - GREEN: Jobs found in both files (312 jobs)
    - LIGHT BLUE: Jobs found only in jobs.md (167 jobs)
    """
    
    # Define colors
    GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    LIGHT_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Light blue
    
    # Green rows - Jobs found in BOTH files
    green_job_ids = [
        # Rhino Roofing Products Limited (26 matches)
        "Rhino July 30 2025", "Rhino -202440 -Aug 6", "203629 - Rhino Aug 6", "203610-Rhino-Aug 7",
        "202432-Rhino-Aug 7", "203627-Rhino-Aug 7", "203639-Rhino-Aug 7", "203685-Rhino-Aug 7",
        "Rhino - Aug - 11", "Rhino - 11 - Aug", "Rhino - 11 - Aug 2025", "Rhino - Aug - 13",
        "Rhino - 13 - Aug", "Rhino-Aug 13-2025", "Rhino-August 14-2025", "Rhino - Aug -14",
        
        # LANWA Sanstha Cement Corporation (25 matches)
        "LAN0214-LAN0231 Lanwa 7 Aug", "LAN0222 - Lanwa Aug 5", "LAN0226 - Lanwa Aug 5",
        "LAN0234 - Lsnwa Aug 5", "LAN0235 - Lanwa Aug -5", "Lanwa - 5 Aug", "Lanwa - 11 Aug",
        "Lanwa - 11 - Aug", "Lanwa - Aug - 11", "Lanwa - Aug - 12", "Lanwa - Aug - 14",
        "TO023876 - Lanwa - 7 Aug", "TO024305 - Lanwa Aug 6",
        
        # Sumithra Hasalaka Pvt Ltd (27 matches)
        "SH10058-July", "SH10060-July-2025", "SH10061-July-2025", "SH10062-July-2025",
        "SH10063-July-2025", "SH10064-July-2025", "SH10065-July-2025", "SH10066-July-2025",
        "SH10067-July-2025", "SH10068-July-2025", "SH10069-July-2025", "SH10070-July-2025",
        "SH10071-July-2025", "SH10072-July-2025", "SH10073-July-2025", "SH10074-July-2025",
        "SH10075-July-2025", "SH10076-August-2025", "SH10077-August-2025", "SH10078-August-2025",
        "SH10079-August-2025", "SH10080-August-2025", "SH10081-August-2025", "SH10082-August-2025",
        "SH10083 New-August-2025", "SH10084-August-2025", "SH10086-July-2025",
        
        # Sumithra Garments Private Limited (69 matches)
        "SP-July-2025", "SP10056-July-2025", "SP10057-July-2025", "SP10058-July-2025",
        "SP10059-July-2025", "SP10066-July-2025", "SP10067-July-2025", "SP10068-July-2025",
        "SP10070-July-2025", "SP10071-August-2025", "SP10072-July-2025", "SP10073-August-2025",
        "SP10074-August-2025", "SP10075-August-2025", "SP10077-August-2025", "SP10078New-August-2025",
        "SP10079-August-2025", "SP10080-August-2025", "SP10081-August-2025", "SW10042-July-2025",
        "SW10042New-July-2025", "SW10054", "SW10055-July-2025", "SW10056-July-2025",
        "SW10057-July-2025", "SW10058-July-2025", "SW10059-July-2025", "SW10060-July-2025",
        "SW10061 N-July-2025", "SW10062-August-2025", "SW10063-August-2025", "SW10064-August-2025",
        "SW10065-August-2025", "SW10067-August-2025", "SW10068-August-2025", "SWA10076",
        "SWA10077-July-2025", "SWA10078-July-2025", "SWA10079-July-2025", "SWA10080-July-2025",
        "SWA10081-July-2025", "SWA10082-July-2025", "SWA10083-July-2025", "SWA10084-July-2025",
        "SWA10092-July-2025", "SWA10093-July-2025", "SWA10094-July-2025", "SWA10098-July-2025",
        "SWA10099-July-2025", "SWA10104-July-2025", "SWA10109-Balance Commitment KM(LM-8812)-July-2025",
        "SWA10110-Balance Commitment KM(47-0568)-July-2025", "SWA10113-August-2025",
        "SWA10114-August-2025", "SWA10115-August-2025", "SWA10116-August-2025", "SWA10117-August-2025",
        "SWA10118-August-2025", "SWA10119-August-2025", "SWA10120-August-2025", "SWA10121-August-2025",
        "SWA10122 New-August-2025", "SWA10123-August-2025", "SWA10124-August-2025",
        "SWA10125 New-August-2025", "SWA10126-August-2025", "SWA10127-July-2025", "SWA10128-August-2025",
        
        # Causeway Paints Lanka (125 matches) - Numeric codes
        "1010218075", "1010221770", "1010224717", "1010233314", "1010235905", "1010237043",
        "1010237595", "1010237830", "1010254728", "1010257283", "1010263169", "1010263549",
        "1010266340", "1010266341", "1010266345", "1010266935", "1010266937", "1010267161",
        "1010268864", "1010268910", "1010269021", "1010269028", "1010269643", "1010269928",
        "1010269949", "101027024", "1010270824", "1010271640", "1010271724",
        
        # INSEE Ecocycle Lanka (6 matches)
        "Insee July 28 2025", "Insee July 29 2025", "Insee July 30 2025", "Insee July 31 2025",
        "Insee August 4 2025", "Insee August 5 2025",
        
        # Cisco Speciality Packaging (111 matches)
        "Cisco", "Kithsiri LM-6923 Cisco trips CS0902", "Kithsiri LM-6923 Cisco trips CS0903",
        "Kithsiri LM-6923 Cisco trips CS0904", "Kithsiri LM-6923 Cisco trips CS0905",
        "Kithsiri LM-6923 Cisco trips CS0906", "Kithsiri LM-6923 Cisco trips CS0907",
        "Kithsiri LM-6923 Cisco trips CS0908", "Kithsiri LM-6923 Cisco trips CS0909",
        "Kithsiri LM-6923 Cisco trips CS0970", "Kithsiri LM-6923 Cisco trips CS0971",
        "Kithsiri LM-6923 Cisco trips CS0972", "Kithsiri LM-6923 Cisco trips CS0973",
        "LB-1882", "LB-1882 Cisco 01.08.2025", "LB-1882 Cisco 04.08.2025",
        "LB-1882 Cisco 06.08.2025", "LB-1882 Cisco 31.07.2025", "LL-2883 Cisco 30.07.2025",
        "LL-2883 Cisco 31.07.2025", "Shelton LI-2901 Cisco trips", "Shelton LI-2901 Cisco trips CS0948",
        "Shelton LI-2901 Cisco trips CS0949", "Shelton LI-2901 Cisco trips CS0951",
        "Shelton LI-2901 Cisco trips CS0954", "Shelton - LI-2901 - Cisco - 01.08.2025",
        
        # Little Lion Associates (32 matches)
        "TR2503-28/02", "WH2507-28/02", "WH2508-01/05",
        
        # Kenilworth International Lanka (3 matches)
        "Pasindu Kenilworth KWD0074", "Pasindu Kenilworth KWD0075", "Pasindu Kenilworth KWD0076",
        
        # Siam City Cement Lanka Ltd - Insee (49 matches)
        "DIMO JULY 25 2025", "DIMO July 25 2025", "INSEE JULY 25 2025", "INSEE July 25 2025",
        "Insee July 25 2025", "Insee July 28 2025", "Insee July 29 2025", "Insee July 30 2025",
        "Insee July 31 2025", "Insee Agust 4 2025", "Insee August 4 2025", "Insee August 5 2025",
        "SA00337", "SA00338", "SA00339", "SA00340", "SA00341", "SA00342"
    ]
    
    # Light blue rows - Jobs found ONLY in jobs.md
    light_blue_job_ids = [
        # Additional Cisco Jobs (not in Haulmatic)
        "Vijith LB-1882 Cisco trips", "Vijith LB-1882 Cisco trips CS0911", "Vijith LB-1882 Cisco trips CS0912",
        "Vijith LB-1882 Cisco trips CS0913", "Vijith LB-1882 Cisco trips CS0914", "Vijith LB-1882 Cisco trips CS0915",
        "Vijith LB-1882 Cisco trips CS0916", "Vijith LB-1882 Cisco trips CS0917", "Vijith LB-1882 Cisco trips CS0918",
        "Vijith LB-1882 Cisco trips CS0919", "Vijith LB-1882 Cisco trips CS0920", "Vijith LB-1882 Cisco trips CS0921",
        "Vijith LL-2883 Cisco trips JULY", "Vijith LL-2883 Cisco trips CS0924", "Vijith LL-2883 Cisco trips CS0925",
        "Vijith LL-2883 Cisco trips CS0926", "Vijith LL-2883 Cisco trips CS0927", "Vijith LL-2883 Cisco trips CS0928",
        "Vijith LL-2883 Cisco trips CS0929", "Vijith LL-2883 Cisco trips CS0930", "Vijith LL-2883 Cisco trips CS0931",
        "Vijith LL-2883 Cisco trips CS0932", "Vijith LL-2883 Cisco trips CS0933", "Vijith LL-2883 Cisco trips CS0934",
        "Vijith LL-2883 Cisco trips CS0935", "Vijith LL-2883 Cisco trips CS0936", "Vijith LL-2883 Cisco trips CS0937",
        "Vijith LL-2883 Cisco trips CS0938", "Vijith LL-2883 Cisco trips CS0939", "Vijith LL-2883 Cisco trips CS0940",
        "Vijith LL-2883 Cisco trips CS0941", "Vijith LL-2883 Cisco trips CS0942", "Vijith LL-2883 Cisco trips CS0943",
        "Vijith LL-2883 Cisco trips CS0944", "Vijith LL-2883 Cisco trips CS0945", "Vijith LL-2883 Cisco trips CS0946",
        "Vijith LL-2883 Cisco trips CS0947", "Vijith LL-2883 Cisco trips CS0948", "Vijith LL-2883 Cisco trips CS0949",
        "Vijith LL-2883 Cisco trips CS0950", "Vijith LL-2883 Cisco trips CS0952", "Vijith LL-2883 Cisco trips CS0953",
        "Vijith LL-2883 Cisco trips CS0955", "Vijith LL-2883 Cisco trips CS0956", "Vijith LL-2883 Cisco trips CS0959",
        "Vijith LL-2883 Cisco trips CS0960", "Vijith LL-2883 Cisco trips CS0961", "Vijith LL-2883 Cisco trips CS0962",
        "Vijith LL-2883 Cisco trips CS0963", "Vijith LL-2883 Cisco trips CS0964", "Vijith LL-2883 Cisco trips CS0966",
        "Vijith LL-2883 Cisco trips CS0967", "Vijith LL-2883 Cisco trips CS0969", "Vijith LL-2883 Cisco trips CS0974",
        "Vijith LL-2883 Cisco trips CS0966", "Vijith LL-2883 Cisco trips CS0977", "Vijith LL-2883 Cisco trips CS0965",
        "Vijith LL-2883 Cisco trips CS0975", "Vijith LL-2883 Cisco trips CS0968", "Vijith LL-2883 Cisco trips CS0958",
        "Vijith LL-2883 Cisco trips CS0978", "Vijith LL-2883 Cisco trips CS0979", "Vijith LL-2883 Cisco trips CS0980",
        "Vijith LL-2883 Cisco trips CS0984", "Vijith LL-2883 Cisco trips CS0985", "Vijith LL-2883 Cisco trips CS0986",
        "Vijith LL-2883 Cisco trips CS0987", "Vijith LL-2883 Cisco trips CS0988", "Vijith LL-2883 Cisco trips CS0989",
        "Vijith LL-2883 Cisco trips CS0990", "Vijith LL-2883 Cisco trips CS0991", "Vijith LL-2883 Cisco trips CS0992",
        "Vijith LL-2883 Cisco trips CS0993", "Vijith LL-2883 - Cisco 31.07.2025", "Vijith LL-2883 - Cisco 01.08.2025",
        "Vijith LL-2883 - Cisco 02.08.2025", "Vijith LL-2883 - Cisco 04.08.2025", "Vijith LL-2883 - Cisco 05.08.2025",
        
        # Shelton LI-2901 Cisco trips
        "Shelton LI-2901 Cisco trips", "Shelton LI-2901 Cisco trips CS0950", "Shelton LI-2901 Cisco trips CS0952",
        "Shelton LI-2901 Cisco trips CS0953", "Shelton LI-2901 Cisco trips CS0955", "Shelton LI-2901 Cisco trips CS0956",
        "Shelton LI-2901 Cisco trips CS0959", "Shelton LI-2901 Cisco trips CS0960", "Shelton LI-2901 Cisco trips CS0961",
        "Shelton LI-2901 Cisco trips CS0962", "Shelton LI-2901 Cisco trips CS0963", "Shelton LI-2901 Cisco trips CS0964",
        "Shelton LI-2901 Cisco trips CS0966", "Shelton LI-2901 Cisco trips CS0967", "Shelton LI-2901 Cisco trips CS0969",
        "Shelton LI-2901 Cisco trips CS0974", "Shelton LI-2901 Cisco trips CS0966", "Shelton LI-2901 Cisco trips CS0977",
        "Shelton LI-2901 Cisco trips CS0965", "Shelton LI-2901 Cisco trips CS0975", "Shelton LI-2901 Cisco trips CS0968",
        "Shelton LI-2901 Cisco trips CS0958", "Shelton LI-2901 Cisco trips CS0978", "Shelton LI-2901 Cisco trips CS0979",
        "Shelton LI-2901 Cisco trips CS0980", "Shelton LI-2901 Cisco trips CS0984", "Shelton LI-2901 Cisco trips CS0985",
        "Shelton LI-2901 Cisco trips CS0986", "Shelton LI-2901 Cisco trips CS0987", "Shelton LI-2901 Cisco trips CS0988",
        "Shelton LI-2901 Cisco trips CS0989", "Shelton LI-2901 Cisco trips CS0990", "Shelton LI-2901 Cisco trips CS0991",
        "Shelton LI-2901 Cisco trips CS0992", "Shelton LI-2901 Cisco trips CS0993", "Shelton - LI-2901 - Cisco - 31.07.2025",
        "Shelton - LI-2901 - Cisco - 02.08.2025", "Shelton - LI-2901 - Cisco - 04.08.2025", "Shelton - LI-2901 - Cisco - 05.08.2025",
        "Shelton - LI-2901 - Cisco - 06.08.2025", "Shelton - LI-2901 - Cisco - 07.08.2025",
        
        # Additional Little Lion Jobs (not in Haulmatic)
        "TR2507-23/02", "TR2507-29/03", "TR2507-23/05", "TR2508-01/01", "TR2507-29/02",
        "TR2508-05/04", "TR2507-31/01", "TR2507-31/02", "TR2508-06/03", "TR2508-06/02",
        "TR2508-06/04", "TR2508-07/02", "TR2508-07/04", "TR2508-07/06", "TR2508-07/08",
        "TR2508-09/01", "TR2508-05/01", "TR2508-05/03", "TR2508-05/02", "TR2508-07/01",
        "TR2508-07/03", "TR2508-07/05", "TR2508-07/07",
        
        # Additional Miscellaneous Jobs (not in Haulmatic)
        "Causeway_28.07_Payments", "PQ 2456", "SH-July-2025", "Eco July 30 2025",
        "1010286770", "1010288811", "1010284216", "1010283766", "1010285084", "1010283059",
        "1010283208", "1010280373", "1010291318", "1010289406", "1010286051", "1010286579",
        "1010288690", "1010296098", "1010305528", "1010302855", "1010301777", "1010298847",
        "1010297096", "1010295454"
    ]
    
    try:
        # Load the workbook
        print("Loading sample_jobs.xlsx...")
        workbook = openpyxl.load_workbook('sample_jobs.xlsx')
        
        # Get the active sheet (or first sheet)
        sheet = workbook.active
        print(f"Working with sheet: {sheet.title}")
        
        # Find the column that contains job IDs
        job_id_column = None
        header_row = 1
        
        # Look for common job ID column headers
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value and any(keyword in str(cell_value).lower() for keyword in ['job', 'id', 'code', 'reference', 'order']):
                job_id_column = col
                print(f"Found job ID column: {col} - '{cell_value}'")
                break
        
        if job_id_column is None:
            # If no specific header found, try to find by content pattern
            for col in range(1, sheet.max_column + 1):
                for row in range(2, min(10, sheet.max_row + 1)):  # Check first few rows
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and any(pattern in str(cell_value) for pattern in ['Rhino', 'LAN', 'SH', 'SP', 'SW', 'SWA', '1010', 'Cisco', 'TR', 'WH', 'KWD', 'SA00', 'Insee']):
                        job_id_column = col
                        print(f"Found job ID column by content pattern: {col}")
                        break
                if job_id_column:
                    break
        
        if job_id_column is None:
            print("Could not identify job ID column. Please specify the column number manually.")
            return
        
        # Process rows and apply colors
        green_count = 0
        light_blue_count = 0
        total_rows = 0
        
        print(f"Processing rows in column {job_id_column}...")
        
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            cell_value = sheet.cell(row=row, column=job_id_column).value
            if cell_value:
                total_rows += 1
                cell_value_str = str(cell_value).strip()
                
                # Check if this job ID matches any in our lists
                if any(job_id.strip() == cell_value_str for job_id in green_job_ids):
                    # Apply green color to entire row
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = GREEN_FILL
                    green_count += 1
                    print(f"Row {row}: Applied GREEN to '{cell_value_str}'")
                    
                elif any(job_id.strip() == cell_value_str for job_id in light_blue_job_ids):
                    # Apply light blue color to entire row
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = LIGHT_BLUE_FILL
                    light_blue_count += 1
                    print(f"Row {row}: Applied LIGHT BLUE to '{cell_value_str}'")
        
        # Save the workbook
        output_filename = 'jobs_colored.xlsx'
        workbook.save(output_filename)
        print(f"\nColoring completed!")
        print(f"Total rows processed: {total_rows}")
        print(f"Green rows (both files): {green_count}")
        print(f"Light blue rows (jobs.md only): {light_blue_count}")
        print(f"Output saved as: {output_filename}")
        
    except FileNotFoundError:
        print("Error: sample_jobs.xlsx file not found in the current directory.")
    except Exception as e:
        print(f"Error processing file: {str(e)}")

if __name__ == "__main__":
    color_job_rows()
