import openpyxl
from collections import defaultdict

def generate_summary_report():
    """Generate a detailed summary report of the colored rows"""
    
    try:
        # Load the colored workbook
        workbook = openpyxl.load_workbook('jobs_colored_corrected.xlsx')
        sheet = workbook.active
        
        print("=" * 80)
        print("EXCEL FILE COLORING SUMMARY REPORT")
        print("=" * 80)
        print(f"File: jobs_colored_corrected.xlsx")
        print(f"Sheet: {sheet.title}")
        print(f"Total Rows: {sheet.max_row}")
        print("=" * 80)
        
        # Analyze colored rows
        green_jobs = []
        light_blue_jobs = []
        uncolored_jobs = []
        
        for row in range(2, sheet.max_row + 1):
            job_id = sheet.cell(row=row, column=1).value
            company_name = sheet.cell(row=row, column=3).value
            row_fill = sheet.cell(row=row, column=1).fill
            
            if row_fill.start_color.rgb == 'FF90EE90':  # Green
                green_jobs.append((row, job_id, company_name))
            elif row_fill.start_color.rgb == 'FF87CEEB':  # Light Blue
                light_blue_jobs.append((row, job_id, company_name))
            else:
                uncolored_jobs.append((row, job_id, company_name))
        
        # Print summary
        print(f"\n🟢 GREEN ROWS (Jobs Found in BOTH Files): {len(green_jobs)}")
        print("-" * 60)
        
        # Group green jobs by company
        green_by_company = defaultdict(list)
        for row, job_id, company in green_jobs:
            green_by_company[company].append((row, job_id))
        
        for company, jobs in green_by_company.items():
            print(f"\n{company} ({len(jobs)} matches):")
            for row, job_id in jobs[:10]:  # Show first 10
                print(f"  Row {row}: {job_id}")
            if len(jobs) > 10:
                print(f"  ... and {len(jobs) - 10} more")
        
        print(f"\n🔵 LIGHT BLUE ROWS (Jobs Found ONLY in jobs.md): {len(light_blue_jobs)}")
        print("-" * 60)
        
        # Group light blue jobs by company
        light_blue_by_company = defaultdict(list)
        for row, job_id, company in light_blue_jobs:
            light_blue_by_company[company].append((row, job_id))
        
        for company, jobs in light_blue_by_company.items():
            print(f"\n{company} ({len(jobs)} matches):")
            for row, job_id in jobs:
                print(f"  Row {row}: {job_id}")
        
        print(f"\n⚪ UNCOLORED ROWS: {len(uncolored_jobs)}")
        print("-" * 60)
        
        # Show some uncolored examples
        uncolored_by_company = defaultdict(list)
        for row, job_id, company in uncolored_jobs:
            uncolored_by_company[company].append((row, job_id))
        
        for company, jobs in list(uncolored_by_company.items())[:5]:  # Show first 5 companies
            print(f"\n{company} ({len(jobs)} uncolored):")
            for row, job_id in jobs[:3]:  # Show first 3
                print(f"  Row {row}: {job_id}")
            if len(jobs) > 3:
                print(f"  ... and {len(jobs) - 3} more")
        
        if len(uncolored_by_company) > 5:
            print(f"\n... and {len(uncolored_by_company) - 5} more companies with uncolored rows")
        
        print("\n" + "=" * 80)
        print("SUMMARY:")
        print(f"✅ Green rows: {len(green_jobs)}")
        print(f"🔵 Light blue rows: {len(light_blue_jobs)}")
        print(f"⚪ Uncolored rows: {len(uncolored_jobs)}")
        print(f"📊 Total processed: {len(green_jobs) + len(light_blue_jobs) + len(uncolored_jobs)}")
        print("=" * 80)
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")

if __name__ == "__main__":
    generate_summary_report()
