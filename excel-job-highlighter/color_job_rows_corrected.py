import openpyxl
from openpyxl.styles import PatternFill
import re

def color_job_rows_corrected():
    """
    Color code rows in Excel files based on company names in the jobname column
    """
    
    # Define colors
    GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    
    # Green rows - Jobs found in BOTH files (company names)
    green_companies = [
        "Rhino", "Lanwa", "Sumithra Hasalaka", "Sumithra Garment", "Causeway", "Insee", "Cisco", 
        "Little Lion", "Kenilworth", "Siam city"
    ]
    
    # Light blue rows - Jobs found ONLY in jobs.md (specific patterns)
    light_blue_patterns = [
        r'Vijith.*Cisco', r'Shelton.*Cisco.*CS\d+', r'TR2507.*\d+/\d+', r'TR2508.*\d+/\d+',
        r'Causeway_28\.07_Payments', r'PQ 2456', r'SH-July-2025', r'Eco July 30 2025'
    ]
    
    try:
        # Load the workbook
        print("Loading sample_jobs.xlsx...")
        workbook = openpyxl.load_workbook('sample_jobs.xlsx')
        sheet = workbook.active
        print(f"Working with sheet: {sheet.title}")
        
        # Company name is in column 3 (jobname)
        company_column = 3
        print(f"Using company name column: {company_column} - '{sheet.cell(row=1, column=company_column).value}'")
        
        # Process rows
        green_count = 0
        light_blue_count = 0
        total_rows = 0
        
        print(f"Processing rows based on company names...")
        
        for row in range(2, sheet.max_row + 1):
            company_name = sheet.cell(row=row, column=company_column).value
            if company_name:
                total_rows += 1
                company_name_str = str(company_name).strip()
                
                # Check if company name matches green list
                is_green = any(company.lower() in company_name_str.lower() for company in green_companies)
                
                # Check if job name matches light blue patterns
                is_light_blue = any(re.search(pattern, company_name_str, re.IGNORECASE) for pattern in light_blue_patterns)
                
                if is_green:
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = GREEN_FILL
                    green_count += 1
                    print(f"Row {row}: GREEN - '{company_name_str}'")
                    
                elif is_light_blue:
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = LIGHT_BLUE_FILL
                    light_blue_count += 1
                    print(f"Row {row}: LIGHT BLUE - '{company_name_str}'")
        
        # Save
        output_filename = 'jobs_colored_corrected.xlsx'
        workbook.save(output_filename)
        print(f"\nColoring completed!")
        print(f"Total rows processed: {total_rows}")
        print(f"Green rows (both files): {green_count}")
        print(f"Light blue rows (jobs.md only): {light_blue_count}")
        print(f"Output saved as: {output_filename}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    color_job_rows_corrected()
