import openpyxl

def examine_excel():
    """Examine the content of Excel files to understand the structure"""
    
    try:
        workbook = openpyxl.load_workbook('sample_jobs.xlsx')
        sheet = workbook.active
        
        print(f"Sheet: {sheet.title}")
        print(f"Rows: {sheet.max_row}")
        print(f"Columns: {sheet.max_column}")
        print("\nFirst 10 rows of data:")
        print("-" * 50)
        
        for row in range(1, min(11, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(6, sheet.max_column + 1)):  # First 5 columns
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "None")
            print(f"Row {row}: {' | '.join(row_data)}")
        
        print("\nSample job IDs from first 20 rows:")
        print("-" * 50)
        
        for row in range(2, min(22, sheet.max_row + 1)):
            job_id = sheet.cell(row=row, column=1).value
            if job_id:
                print(f"Row {row}: '{job_id}' (Type: {type(job_id)})")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    examine_excel()
