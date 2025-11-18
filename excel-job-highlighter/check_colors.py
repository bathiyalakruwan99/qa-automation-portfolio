import openpyxl

def check_colors():
    """Check the actual color values in the Excel file"""
    
    try:
        workbook = openpyxl.load_workbook('jobs_colored_corrected.xlsx')
        sheet = workbook.active
        
        print("Checking colors in the Excel file...")
        print("=" * 50)
        
        # Check first few rows
        for row in range(2, 10):
            cell = sheet.cell(row=row, column=1)
            fill = cell.fill
            print(f"Row {row}: Fill type: {fill.fill_type}")
            print(f"  Start color: {fill.start_color.rgb if fill.start_color.rgb else 'None'}")
            print(f"  End color: {fill.end_color.rgb if fill.end_color.rgb else 'None'}")
            print(f"  Pattern type: {fill.patternType}")
            print()
        
        # Check some specific rows that should be colored
        test_rows = [2, 3, 4, 5, 52, 53, 55, 56, 214, 215]
        print("Checking specific test rows:")
        print("-" * 30)
        
        for row in test_rows:
            if row <= sheet.max_row:
                cell = sheet.cell(row=row, column=1)
                fill = cell.fill
                company = sheet.cell(row=row, column=3).value
                print(f"Row {row} ({company}):")
                print(f"  Start color: {fill.start_color.rgb if fill.start_color.rgb else 'None'}")
                print(f"  End color: {fill.end_color.rgb if fill.end_color.rgb else 'None'}")
                print(f"  Pattern type: {fill.patternType}")
                print()
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    check_colors()
