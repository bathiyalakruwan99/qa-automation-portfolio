import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime

class ExcelCorrector:
    def __init__(self):
        # Initialize tracking for all changes with detailed before/after
        self.hr_red_cell_changes = []
        self.state_changes = []
        self.division_corrections = []
        self.principle_contact_changes = []
        
        # New: Detailed change tracking for visual reporting
        self.detailed_changes = {
            'organization': [],
            'divisions': [],
            'human_resources': [],
            'vehicles': [],
            'locations': []
        }
        
        # State name mapping for corrections - updated to use district format
        self.state_corrections = {
            'western': 'Colombo District',
            'central': 'Kandy District',
            'southern': 'Galle District',
            'northern': 'Jaffna District',
            'eastern': 'Batticaloa District',
            'north western': 'Kurunegala District',
            'north central': 'Anuradhapura District',
            'uva': 'Badulla District',
            'sabaragamuwa': 'Ratnapura District',
            'colombo': 'Colombo District',
            'gampaha': 'Gampaha District',
            'kalutara': 'Kalutara District',
            'kandy': 'Kandy District',
            'matale': 'Matale District',
            'nuwara eliya': 'Nuwara Eliya District',
            'galle': 'Galle District',
            'matara': 'Matara District',
            'hambantota': 'Hambantota District',
            'jaffna': 'Jaffna District',
            'kilinochchi': 'Kilinochchi District',
            'mannar': 'Mannar District',
            'vavuniya': 'Vavuniya District',
            'mullaitivu': 'Mullaitivu District',
            'batticaloa': 'Batticaloa District',
            'ampara': 'Ampara District',
            'trincomalee': 'Trincomalee District',
            'kurunegala': 'Kurunegala District',
            'puttalam': 'Puttalam District',
            'anuradhapura': 'Anuradhapura District',
            'polonnaruwa': 'Polonnaruwa District',
            'badulla': 'Badulla District',
            'moneragala': 'Monaragala District',
            'ratnapura': 'Ratnapura District',
            'kegalle': 'Kegalle District'
        }
    
    def correct_state_name(self, state_value, org_name=None):
        """Convert state name to correct Sri Lankan district format"""
        if pd.isna(state_value) or state_value == '':
            return 'Gampaha District'  # Default
        
        original_state = str(state_value).strip()
        state_lower = original_state.lower().strip()
        
        # Check if it already ends with "district"
        if state_lower.endswith('district'):
            corrected_state = original_state
        else:
            corrected_state = self.state_corrections.get(state_lower, f"{state_value} District")
        
        # Log the change if there was a correction
        if original_state != corrected_state:
            self.log_state_change(original_state, corrected_state, org_name)
        
        return corrected_state
    
    def is_valid_email(self, email):
        """Check if email format is valid"""
        if pd.isna(email) or email == '':
            return False
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(email_pattern, str(email)) is not None
    
    def generate_org_short_name(self, org_name):
        """Generate short name from organization name"""
        if pd.isna(org_name) or org_name == '':
            return 'ORG'
        
        # Use the full organization name, cleaned up
        org_str = str(org_name).strip()
        # Remove special characters and spaces, keep alphanumeric only
        clean_name = ''.join(char for char in org_str if char.isalnum())
        
        # If the cleaned name is too long (over 20 chars), use a shorter version
        if len(clean_name) > 20:
            # Take first 15 chars + last 5 chars to keep it meaningful but not too long
            return clean_name[:15] + clean_name[-5:]
        
        return clean_name if clean_name else 'ORG'
    
    def format_vehicle_category(self, category_value):
        """Format vehicle category to proper format"""
        if pd.isna(category_value) or category_value == '':
            return category_value
        
        category_str = str(category_value).strip()
        
        # Extract number from various formats like "20 FT", "(20 FT)", "20FT", etc.
        import re
        number_match = re.search(r'(\d+)', category_str)
        if number_match:
            number = number_match.group(1)
            return f"{number}Ft"
        else:
            # If no number found, return as is
            return category_value
    
    def is_red_cell(self, cell):
        """Check if a cell has red background color"""
        try:
            if cell.fill and hasattr(cell.fill, 'start_color'):
                if cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                    rgb = cell.fill.start_color.rgb
                    if isinstance(rgb, str) and len(rgb) >= 6:
                        # Handle both ARGB (8 chars) and RGB (6 chars) formats
                        if len(rgb) == 8:  # ARGB format
                            red_component = int(rgb[2:4], 16)
                            green_component = int(rgb[4:6], 16)
                            blue_component = int(rgb[6:8], 16)
                        else:  # RGB format
                            red_component = int(rgb[0:2], 16)
                            green_component = int(rgb[2:4], 16)
                            blue_component = int(rgb[4:6], 16)
                        
                        # Check if it's predominantly red (red > 200 and others < 100)
                        return red_component > 200 and green_component < 100 and blue_component < 100
        except Exception:
            # If any error occurs in color checking, assume it's not red
            pass
        
        return False
    
    def reset_change_tracking(self):
        """Reset all change tracking for a new file"""
        self.hr_red_cell_changes = []
        self.state_changes = []
        self.division_corrections = []
        self.principle_contact_changes = []
        
        # Reset detailed changes tracking
        self.detailed_changes = {
            'organization': [],
            'divisions': [],
            'human_resources': [],
            'vehicles': [],
            'locations': []
        }
    
    def is_end_row(self, df, idx):
        """Check if a row is an END row that should be skipped"""
        try:
            for col in df.columns:
                cell_value = str(df.loc[idx, col]).upper().strip()
                if cell_value == 'END':
                    return True
            return False
        except:
            return False
    
    def log_hr_red_cell_change(self, change_type, row_idx, org_name=None, original_value=None, new_value=None):
        """Log HR red cell changes for reporting"""
        self.hr_red_cell_changes.append({
            'organization': org_name if org_name else 'Unknown',
            'change_type': change_type,  # 'Both Red - Update', 'NIC Red - Prefix', 'Email Red - Prefix'
            'row_number': row_idx + 4,  # Convert to Excel row number
            'original_value': original_value,
            'new_value': new_value
        })
    
    def log_state_change(self, original, corrected, org_name=None):
        """Log state name changes for reporting"""
        if original != corrected:
            self.state_changes.append({
                'organization': org_name if org_name else 'Unknown',
                'original_state': original,
                'corrected_state': corrected,
                'changed': True
            })
    
    def log_division_correction(self, correction_type, org_name=None, original_value=None, new_value=None):
        """Log division corrections for reporting"""
        self.division_corrections.append({
            'organization': org_name if org_name else 'Unknown',
            'correction_type': correction_type,  # 'Division Name', 'Purpose', etc.
            'original_value': original_value if original_value else 'Empty',
            'corrected_value': new_value,
            'changed': original_value != new_value if original_value else True
        })
    
    def log_principle_contact_change(self, field_type, org_name=None, sheet_type="Organization Details"):
        """Log principle contact name changes for reporting"""
        self.principle_contact_changes.append({
            'organization': org_name if org_name else 'Unknown',
            'field_type': field_type,  # 'First Name' or 'Last Name'
            'original_value': 'Empty',
            'filled_value': f"Principle Contact's {field_type}" if field_type else 'Demo Value',
            'sheet': sheet_type
        })
    
    def generate_hr_red_cell_report(self):
        """Generate a report of HR red cell changes"""
        if not self.hr_red_cell_changes:
            return "No HR red cell changes were made."
        
        report = "\n" + "="*80 + "\n"
        report += "            HR RED CELL DETECTION REPORT\n"
        report += "="*80 + "\n\n"
        
        # Summary statistics
        total_changes = len(self.hr_red_cell_changes)
        both_red_no_change = sum(1 for change in self.hr_red_cell_changes if change['change_type'] == 'Both Red - No Change')
        nic_prefix_changes = sum(1 for change in self.hr_red_cell_changes if change['change_type'] == 'NIC Red - Prefix')
        email_prefix_changes = sum(1 for change in self.hr_red_cell_changes if change['change_type'] == 'Email Red - Prefix')
        dummy_nics = sum(1 for change in self.hr_red_cell_changes if change['change_type'] == 'Dummy NIC Generated')
        dummy_emails = sum(1 for change in self.hr_red_cell_changes if change['change_type'] == 'Dummy Email Generated')
        
        report += f"SUMMARY:\n"
        report += f"  Total HR changes: {total_changes}\n"
        report += f"  Dummy NICs generated: {dummy_nics}\n"
        report += f"  Dummy Emails generated: {dummy_emails}\n"
        report += f"  Both NIC & Email red → No changes made: {both_red_no_change}\n"
        report += f"  Only NIC red → Prefix added: {nic_prefix_changes}\n"
        report += f"  Only Email red → Prefix added: {email_prefix_changes}\n\n"
        
        # Group by change type
        dummy_nic_changes = [c for c in self.hr_red_cell_changes if c['change_type'] == 'Dummy NIC Generated']
        dummy_email_changes = [c for c in self.hr_red_cell_changes if c['change_type'] == 'Dummy Email Generated']
        both_red_changes = [c for c in self.hr_red_cell_changes if c['change_type'] == 'Both Red - No Change']
        nic_prefix_changes_list = [c for c in self.hr_red_cell_changes if c['change_type'] == 'NIC Red - Prefix']
        email_prefix_changes_list = [c for c in self.hr_red_cell_changes if c['change_type'] == 'Email Red - Prefix']
        
        # Dummy NIC generation
        if dummy_nic_changes:
            report += "DUMMY NICs GENERATED:\n"
            report += "-" * 60 + "\n"
            for change in dummy_nic_changes:
                report += f"  Organization: {change['organization']}\n"
                report += f"  Row: {change['row_number']} (Excel row)\n"
                report += f"  Original: '{change['original_value']}'\n"
                report += f"  Generated: '{change['new_value']}'\n\n"
        
        # Dummy Email generation
        if dummy_email_changes:
            report += "DUMMY EMAILS GENERATED:\n"
            report += "-" * 60 + "\n"
            for change in dummy_email_changes:
                report += f"  Organization: {change['organization']}\n"
                report += f"  Row: {change['row_number']} (Excel row)\n"
                report += f"  Original: '{change['original_value']}'\n"
                report += f"  Generated: '{change['new_value']}'\n\n"
        
        # Both red changes (No changes made)
        if both_red_changes:
            report += "BOTH NIC & EMAIL RED → NO CHANGES MADE:\n"
            report += "-" * 60 + "\n"
            for change in both_red_changes:
                report += f"  Organization: {change['organization']}\n"
                report += f"  Row: {change['row_number']} (Excel row)\n"
                report += f"  Action: No changes made\n"
                report += f"  Reason: Both NIC and Email cells are red (existing in system)\n\n"
        
        # NIC prefix changes
        if nic_prefix_changes_list:
            report += "ONLY NIC RED → PREFIX ADDED:\n"
            report += "-" * 60 + "\n"
            for change in nic_prefix_changes_list:
                report += f"  Organization: {change['organization']}\n"
                report += f"  Row: {change['row_number']} (Excel row)\n"
                report += f"  Original NIC: '{change['original_value']}'\n"
                report += f"  New NIC: '{change['new_value']}'\n\n"
        
        # Email prefix changes  
        if email_prefix_changes_list:
            report += "ONLY EMAIL RED → PREFIX ADDED:\n"
            report += "-" * 60 + "\n"
            for change in email_prefix_changes_list:
                report += f"  Organization: {change['organization']}\n"
                report += f"  Row: {change['row_number']} (Excel row)\n"
                report += f"  Original Email: '{change['original_value']}'\n"
                report += f"  New Email: '{change['new_value']}'\n\n"
        
        report += "RECOMMENDATIONS:\n"
        report += "-" * 40 + "\n"
        report += "• Review red cell changes for accuracy\n"
        report += "• Both red cells indicate existing records - no changes needed\n"
        report += "• Check that prefixed values maintain data integrity\n"
        report += "• Confirm red cells indicate existing system records\n\n"
        
        report += "="*80 + "\n"
        return report
    
    def generate_comprehensive_report(self):
        """Generate comprehensive report including all corrections and changes organized by sheet"""
        comprehensive_report = ""
        
        # Header
        comprehensive_report += "="*80 + "\n"
        comprehensive_report += "         COMPLETE FILE CORRECTION REPORT - VISUAL CHANGES\n"
        comprehensive_report += "="*80 + "\n\n"
        
        # Processing Summary at the top
        total_tracked_changes = (len(self.hr_red_cell_changes) + len(self.state_changes) + 
                               len(self.division_corrections) + len(self.principle_contact_changes))
        
        # Calculate total detailed changes
        total_detailed_changes = sum(len(changes) for changes in self.detailed_changes.values())
        
        comprehensive_report += "PROCESSING OVERVIEW:\n"
        comprehensive_report += "-" * 40 + "\n"
        comprehensive_report += f"✅ File processed successfully\n"
        comprehensive_report += f"✅ 5 sheets processed: Organization, Divisions, Human Resources, Vehicles, Locations\n"
        comprehensive_report += f"✅ Total visual changes made: {total_detailed_changes}\n"
        comprehensive_report += f"✅ File ready for bulk upload\n\n"
        
        # Generate detailed changes report for each sheet
        sheet_names = {
            'organization': 'ORGANIZATION DETAILS',
            'divisions': 'DIVISIONS', 
            'human_resources': 'HUMAN RESOURCES',
            'vehicles': 'VEHICLES',
            'locations': 'LOCATIONS'
        }
        
        for sheet_key, sheet_display_name in sheet_names.items():
            comprehensive_report += "="*80 + "\n"
            comprehensive_report += f"                    SHEET: {sheet_display_name}\n"
            comprehensive_report += "="*80 + "\n\n"
            
            changes = self.detailed_changes.get(sheet_key, [])
            
            if changes:
                comprehensive_report += f"DETAILED CHANGES MADE ({len(changes)} changes):\n"
                comprehensive_report += "-" * 60 + "\n"
                
                # Group changes by type
                changes_by_type = {}
                for change in changes:
                    change_type = change['change_type']
                    if change_type not in changes_by_type:
                        changes_by_type[change_type] = []
                    changes_by_type[change_type].append(change)
                
                # Display each change type
                for change_type, type_changes in changes_by_type.items():
                    comprehensive_report += f"\n📝 {change_type} ({len(type_changes)} changes):\n"
                    comprehensive_report += "    " + "-" * 50 + "\n"
                    
                    for change in type_changes:
                        comprehensive_report += f"    • {change['organization']} - {change['row_info']}\n"
                        comprehensive_report += f"      Field: {change['field_name']}\n"
                        comprehensive_report += f"      BEFORE: '{change['original_value']}'\n"
                        comprehensive_report += f"      AFTER:  '{change['new_value']}'\n"
                        comprehensive_report += f"      ──────────────────────────────────────\n"
                
                comprehensive_report += "\n"
            else:
                comprehensive_report += "NO CUSTOM CHANGES REQUIRED\n"
                comprehensive_report += "✅ All standard corrections applied automatically\n\n"
        
        # Add legacy reporting for HR red cells if present
        if self.hr_red_cell_changes:
            comprehensive_report += "="*80 + "\n"
            comprehensive_report += "                    HR RED CELL DETAILED REPORT\n"
            comprehensive_report += "="*80 + "\n\n"
            
            for change in self.hr_red_cell_changes:
                comprehensive_report += f"📍 {change['organization']} - Row {change['row_number']}\n"
                comprehensive_report += f"   Change Type: {change['change_type']}\n"
                if 'original_value' in change:
                    comprehensive_report += f"   Original: '{change['original_value']}'\n"
                if 'new_value' in change:
                    comprehensive_report += f"   New: '{change['new_value']}'\n"
                comprehensive_report += "\n"
        
        comprehensive_report += "="*80 + "\n"
        comprehensive_report += "                      FINAL PROCESSING SUMMARY\n"
        comprehensive_report += "="*80 + "\n"
        comprehensive_report += f"🎯 TOTAL CHANGES MADE: {total_detailed_changes} detailed corrections\n"
        comprehensive_report += f"📊 SHEETS PROCESSED: All 5 sheets fully corrected\n"
        comprehensive_report += f"✅ RESULT: File ready for bulk upload with all formatting preserved!\n"
        comprehensive_report += "="*80 + "\n"
        
        return comprehensive_report
    
    def get_processing_stats(self):
        """Return basic processing statistics for reports"""
        custom_corrections = len(self.hr_red_cell_changes) + len(self.state_changes) + len(self.division_corrections) + len(self.principle_contact_changes)
        standard_corrections = 16  # Count of standard corrections always applied
        total_corrections = custom_corrections + standard_corrections
        
        return {
            'sheets_processed': '5 (Organization, HR, Divisions, Vehicles, Locations)',
            'total_corrections': f'{total_corrections} (Standard: {standard_corrections}, Custom: {custom_corrections})',
            'processing_time': 'Completed Successfully',
            'file_size': 'Optimized & Ready',
            'input_file': 'Selected Excel File',
            'output_file': 'Corrected Excel File'
        }
    
    def get_detailed_stats(self):
        """Return detailed statistics for comprehensive reporting"""
        total_custom_corrections = (len(self.hr_red_cell_changes) + len(self.state_changes) + 
                                  len(self.division_corrections) + len(self.principle_contact_changes))
        
        return {
            'corrections_by_category': {
                'HR Red Cell Changes': len(self.hr_red_cell_changes),
                'State Name Corrections': len(self.state_changes),
                'Division Corrections': len(self.division_corrections),
                'Principle Contact Changes': len(self.principle_contact_changes),
                'Organization Details': 1,  # Always applied
                'Vehicle Updates': 1,  # Always applied
                'Location Processing': 1,  # Always applied
                'Status Field Updates': 1  # Always applied
            },
            'standard_corrections': {
                'organization_status_to_non_boi': True,
                'organization_verticals_to_vert_trn': True,
                'organization_country_to_sri_lanka': True,
                'hr_gender_standardized': True,
                'hr_division_to_admin': True,
                'divisions_name_to_admin': True,
                'divisions_purpose_to_pps_stg': True,
                'vehicles_division_to_admin': True,
                'vehicles_type_to_truck': True,
                'vehicles_load_type_to_loads': True,
                'vehicles_categories_formatted': True,
                'locations_empty_ids_filled': True,
                'locations_empty_names_filled': True,
                'locations_duplicates_handled': True,
                'all_status_fields_to_create': True,
                'end_markers_added': True
            },
            'processing_details': {
                'total_custom_corrections': total_custom_corrections,
                'dummy_nics_generated': len([c for c in self.hr_red_cell_changes if c['change_type'] == 'Dummy NIC Generated']),
                'dummy_emails_generated': len([c for c in self.hr_red_cell_changes if c['change_type'] == 'Dummy Email Generated']),
                'both_red_detected': len([c for c in self.hr_red_cell_changes if c['change_type'] == 'Both Red - No Change']),
                'nic_prefixes_added': len([c for c in self.hr_red_cell_changes if c['change_type'] == 'NIC Red - Prefix']),
                'email_prefixes_added': len([c for c in self.hr_red_cell_changes if c['change_type'] == 'Email Red - Prefix']),
                'standard_corrections_applied': 16  # Count of standard corrections
            }
        }
    
    def handle_duplicates_and_empty(self, df, column_name, prefix, org_short_names, is_email=False):
        """Handle duplicates and empty cells in NIC or Email columns with detailed tracking"""
        # Convert column to object type to handle mixed data types
        df[column_name] = df[column_name].astype('object')
        
        seen_values = {}
        dummy_counter = 1
        
        for idx in range(len(df)):
            # Skip END rows
            if self.is_end_row(df, idx):
                continue
                
            current_value = df.loc[idx, column_name]
            org_name = org_short_names[idx] if idx < len(org_short_names) else 'Unknown'
            
            # Handle empty or invalid values
            if (pd.isna(current_value) or str(current_value).strip() == '' or 
                (is_email and ('noemail.com' in str(current_value) or 'noemal.com' in str(current_value)))):
                
                original_value = str(current_value) if pd.notna(current_value) else 'Empty'
                
                if is_email:
                    # Generate dummy email: DUMMY + sequential number + @ + OrgShortName + .com
                    new_value = f"DUMMY{dummy_counter:03d}@{org_short_names[idx]}.com"
                    self.log_detailed_change('Human Resources', 'Dummy Email Generated', f'Row {idx + 4}', 
                                           column_name, original_value, new_value, org_name)
                    self.log_hr_red_cell_change('Dummy Email Generated', idx, org_name, original_value, new_value)
                    print(f"Generated dummy Email for row {idx + 4}: {new_value}")
                else:
                    # Generate dummy NIC: DUMMY + sequential number + OrgShortName
                    new_value = f"DUMMY{dummy_counter:03d}{org_short_names[idx]}"
                    self.log_detailed_change('Human Resources', 'Dummy NIC Generated', f'Row {idx + 4}', 
                                           column_name, original_value, new_value, org_name)
                    self.log_hr_red_cell_change('Dummy NIC Generated', idx, org_name, original_value, new_value)
                    print(f"Generated dummy NIC for row {idx + 4}: {new_value}")
                
                df.loc[idx, column_name] = new_value
                seen_values[new_value] = 1
                dummy_counter += 1
                
            else:
                # Handle duplicates
                value_str = str(current_value).strip()
                if value_str in seen_values:
                    seen_values[value_str] += 1
                    original_value = current_value
                    new_value = f"{value_str}_DUPLICATE_{seen_values[value_str]}"
                    
                    self.log_detailed_change('Human Resources', f'Duplicate {column_name} Handled', f'Row {idx + 4}', 
                                           column_name, original_value, new_value, org_name)
                    df.loc[idx, column_name] = new_value
                    print(f"Handled duplicate {column_name} for row {idx + 4}: {value_str} → {new_value}")
                else:
                    seen_values[value_str] = 1
        
        return df
    
    def correct_organization_details(self, df, processing_options=None):
        """Correct Organization Details sheet with enhanced validation and corrections"""
        print("Correcting Organization Details...")
        
        # Get organization names for reporting
        org_col = None
        for col in df.columns:
            if 'organization' in col.lower() or 'company' in col.lower():
                org_col = col
                break
        
        # Find additional columns for enhanced validation
        org_name_col = None
        org_short_name_col = None
        operations_col = None
        principle_contact_first_name_col = None
        principle_contact_last_name_col = None
        address_line_col = None
        city_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'organization' in col_lower and 'name' in col_lower and 'short' not in col_lower:
                org_name_col = col
            elif 'organization' in col_lower and 'short' in col_lower:
                org_short_name_col = col
            elif 'operation' in col_lower:
                operations_col = col
            elif 'principle contact' in col_lower and 'first name' in col_lower:
                principle_contact_first_name_col = col
            elif 'principle contact' in col_lower and 'last name' in col_lower:
                principle_contact_last_name_col = col
            elif 'address line' in col_lower:
                address_line_col = col
            elif 'city' in col_lower:
                city_col = col
        
        # Track organization short names for duplicate handling
        seen_org_short_names = set()
        short_name_counter = 1
        
        # Valid values for validation
        valid_statuses = ['NON_BOI', 'BOI']
        valid_verticals = ['VERT-CUS', 'VERT-SPO', 'VERT-YO', 'VERT-IM-EX', 'VERT-SHIPPING-LINE', 'VERT-TRN']
        valid_districts = [
            'Colombo District', 'Gampaha District', 'Kalutara District', 'Kandy District', 
            'Matale District', 'Nuwara Eliya District', 'Galle District', 'Matara District', 
            'Hambantota District', 'Jaffna District', 'Kilinochchi District', 'Mannar District', 
            'Vavuniya District', 'Mullaitivu District', 'Batticaloa District', 'Ampara District', 
            'Trincomalee District', 'Kurunegala District', 'Anuradhapura District', 
            'Polonnaruwa District', 'Badulla District', 'Monaragala District', 'Ratnapura District', 
            'Kegalle District'
        ]
        
        # Check processing options to determine what corrections to apply
        apply_corrections = True
        apply_dummy_data = True
        
        if processing_options and 'organization' in processing_options:
            org_options = processing_options['organization']
            # Check if Organization Name options exist and get their values
            if 'Organization Name' in org_options:
                apply_corrections = org_options['Organization Name']['correct']
                apply_dummy_data = org_options['Organization Name']['dummy_data']
        
        # Process each row for enhanced corrections
        for idx in range(len(df)):
            if self.is_end_row(df, idx):
                continue
                
            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
            
            # 1. Fill empty Organization Name if needed
            if org_name_col and apply_dummy_data:
                org_name_value = df.loc[idx, org_name_col]
                if pd.isna(org_name_value) or str(org_name_value).strip() == '':
                    new_org_name = f"Organization_{short_name_counter}"
                    self.log_detailed_change('Organization', 'Organization Name Fill', f'Row {idx + 4}', 
                                           org_name_col, 'Empty', new_org_name, org_name)
                    df.loc[idx, org_name_col] = new_org_name
                    short_name_counter += 1
            
            # 2. Handle Organization Short Name duplicates and empty values
            if org_short_name_col and apply_corrections:
                org_short_name_value = df.loc[idx, org_short_name_col]
                if pd.isna(org_short_name_value) or str(org_short_name_value).strip() == '':
                    # Generate unique short name
                    new_short_name = f"ORG{short_name_counter:03d}"
                    self.log_detailed_change('Organization', 'Organization Short Name Fill', f'Row {idx + 4}', 
                                           org_short_name_col, 'Empty', new_short_name, org_name)
                    df.loc[idx, org_short_name_col] = new_short_name
                    seen_org_short_names.add(new_short_name)
                    short_name_counter += 1
                else:
                    short_name_str = str(org_short_name_value).strip()
                    if short_name_str in seen_org_short_names:
                        # Handle duplicate by adding suffix
                        new_short_name = f"{short_name_str}_DUPLICATE_{len(seen_org_short_names) + 1}"
                        self.log_detailed_change('Organization', 'Organization Short Name Duplicate Fix', f'Row {idx + 4}', 
                                               org_short_name_col, short_name_str, new_short_name, org_name)
                        df.loc[idx, org_short_name_col] = new_short_name
                        seen_org_short_names.add(new_short_name)
                    else:
                        seen_org_short_names.add(short_name_str)
            
            # 3. Fill empty Operations column
            if operations_col and apply_dummy_data:
                operations_value = df.loc[idx, operations_col]
                if pd.isna(operations_value) or str(operations_value).strip() == '':
                    new_operations = "Default"
                    self.log_detailed_change('Organization', 'Operations Fill', f'Row {idx + 4}', 
                                           operations_col, 'Empty', new_operations, org_name)
                    df.loc[idx, operations_col] = new_operations
            
            # 4. Fill empty Principle Contact names
            if principle_contact_first_name_col and apply_dummy_data:
                first_name_value = df.loc[idx, principle_contact_first_name_col]
                if pd.isna(first_name_value) or str(first_name_value).strip() == '':
                    new_first_name = "Principle Contact's First Name"
                    self.log_detailed_change('Organization', 'Principle Contact First Name Fill', f'Row {idx + 4}', 
                                           principle_contact_first_name_col, 'Empty', new_first_name, org_name)
                    df.loc[idx, principle_contact_first_name_col] = new_first_name
            
            if principle_contact_last_name_col and apply_dummy_data:
                last_name_value = df.loc[idx, principle_contact_last_name_col]
                if pd.isna(last_name_value) or str(last_name_value).strip() == '':
                    new_last_name = "Principle Contact's Last Name"
                    self.log_detailed_change('Organization', 'Principle Contact Last Name Fill', f'Row {idx + 4}', 
                                           principle_contact_last_name_col, 'Empty', new_last_name, org_name)
                    df.loc[idx, principle_contact_last_name_col] = new_last_name
            
            # 5. Fill empty Address Line
            if address_line_col and apply_dummy_data:
                address_value = df.loc[idx, address_line_col]
                if pd.isna(address_value) or str(address_value).strip() == '':
                    new_address = f"Address Line {short_name_counter}"
                    self.log_detailed_change('Organization', 'Address Line Fill', f'Row {idx + 4}', 
                                           address_line_col, 'Empty', new_address, org_name)
                    df.loc[idx, address_line_col] = new_address
            
            # 6. Fill empty City
            if city_col and apply_dummy_data:
                city_value = df.loc[idx, city_col]
                if pd.isna(city_value) or str(city_value).strip() == '':
                    new_city = "Colombo"
                    self.log_detailed_change('Organization', 'City Fill', f'Row {idx + 4}', 
                                           city_col, 'Empty', new_city, org_name)
                    df.loc[idx, city_col] = new_city
        
        # Standard corrections with detailed tracking
        for col in df.columns:
            # Status corrections - now supports both NON_BOI and BOI
            if 'status' in col.lower() and apply_corrections:
                for idx in range(len(df)):
                    if self.is_end_row(df, idx):
                        continue
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value):
                        status_str = str(original_value).strip()
                        if status_str not in valid_statuses:
                            # Default to NON_BOI if invalid status
                            new_status = 'NON_BOI'
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            self.log_detailed_change('Organization', 'Status Correction', f'Row {idx + 4}', 
                                                   col, original_value, new_status, org_name)
                            df.loc[idx, col] = new_status
                    else:
                        # Fill empty status with NON_BOI
                        new_status = 'NON_BOI'
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Organization', 'Status Fill', f'Row {idx + 4}', 
                                               col, 'Empty', new_status, org_name)
                        df.loc[idx, col] = new_status
            
            # Verticals corrections - now supports multiple valid values and preserves valid ones
            elif 'vertical' in col.lower():
                for idx in range(len(df)):
                    if self.is_end_row(df, idx):
                        continue
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value):
                        verticals_str = str(original_value).strip()
                        
                        # Handle multiple verticals separated by commas
                        if ',' in verticals_str:
                            vertical_list = [v.strip() for v in verticals_str.split(',')]
                            valid_verticals_found = []
                            invalid_verticals_found = []
                            
                            for vertical in vertical_list:
                                vertical = vertical.strip()
                                if vertical in valid_verticals:
                                    valid_verticals_found.append(vertical)
                                else:
                                    # Check for case-insensitive matches
                                    normalized_input = vertical.upper().replace(' ', '').replace('-', '')
                                    is_valid = False
                                    for valid_vertical in valid_verticals:
                                        normalized_valid = valid_vertical.upper().replace(' ', '').replace('-', '')
                                        if normalized_input == normalized_valid:
                                            valid_verticals_found.append(valid_vertical)
                                            is_valid = True
                                            break
                                    if not is_valid:
                                        invalid_verticals_found.append(vertical)
                            
                            # If we have valid verticals, keep them; if none valid, use default
                            if valid_verticals_found:
                                new_verticals = ', '.join(valid_verticals_found)
                                if invalid_verticals_found:
                                    # Log correction for invalid ones
                                    org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                    self.log_detailed_change('Organization', 'Verticals Correction', f'Row {idx + 4}', 
                                                           col, original_value, new_verticals, org_name)
                                df.loc[idx, col] = new_verticals
                            else:
                                # No valid verticals found, use default
                                new_verticals = 'VERT-TRN'
                                org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                self.log_detailed_change('Organization', 'Verticals Correction', f'Row {idx + 4}', 
                                                       col, original_value, new_verticals, org_name)
                                df.loc[idx, col] = new_verticals
                        else:
                            # Single vertical value
                            is_valid = False
                            if verticals_str in valid_verticals:
                                is_valid = True
                            else:
                                # Check for case-insensitive matches
                                normalized_input = verticals_str.upper().replace(' ', '').replace('-', '')
                                for valid_vertical in valid_verticals:
                                    normalized_valid = valid_vertical.upper().replace(' ', '').replace('-', '')
                                    if normalized_input == normalized_valid:
                                        is_valid = True
                                        break
                            
                            if not is_valid:
                                # Default to VERT-TRN if invalid vertical
                                new_vertical = 'VERT-TRN'
                                org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                self.log_detailed_change('Organization', 'Verticals Correction', f'Row {idx + 4}', 
                                                       col, original_value, new_vertical, org_name)
                                df.loc[idx, col] = new_vertical
                    else:
                        # Fill empty vertical with VERT-TRN
                        new_vertical = 'VERT-TRN'
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Organization', 'Verticals Fill', f'Row {idx + 4}', 
                                               col, 'Empty', new_vertical, org_name)
                        df.loc[idx, col] = new_vertical
            
            # Country corrections
            elif 'country' in col.lower():
                for idx in range(len(df)):
                    if self.is_end_row(df, idx):
                        continue
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value) and str(original_value).strip() != 'Sri Lanka':
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Organization', 'Country Correction', f'Row {idx + 4}', 
                                               col, original_value, 'Sri Lanka', org_name)
                        df.loc[idx, col] = 'Sri Lanka'
                    elif pd.isna(original_value):
                        # Fill empty country
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Organization', 'Country Fill', f'Row {idx + 4}', 
                                               col, 'Empty', 'Sri Lanka', org_name)
                        df.loc[idx, col] = 'Sri Lanka'
            
            # State corrections - now validates against valid district list
            elif 'state' in col.lower() or 'province' in col.lower():
                for idx in range(len(df)):
                    if self.is_end_row(df, idx):
                        continue
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value):
                        state_str = str(original_value).strip()
                        if state_str not in valid_districts:
                            # Try to correct the state name
                            corrected_state = self.correct_state_name(str(original_value), org_name)
                            if corrected_state not in valid_districts:
                                # If still not valid, default to Gampaha District
                                corrected_state = 'Gampaha District'
                            if str(original_value).strip() != corrected_state:
                                org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                self.log_detailed_change('Organization', 'State/Province Correction', f'Row {idx + 4}', 
                                                       col, original_value, corrected_state, org_name)
                                df.loc[idx, col] = corrected_state
                    else:
                        # Fill empty state with default
                        corrected_state = 'Gampaha District'
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Organization', 'State/Province Fill', f'Row {idx + 4}', 
                                               col, 'Empty', corrected_state, org_name)
                        df.loc[idx, col] = corrected_state
        
        return df
    
    def correct_divisions(self, df, processing_options=None):
        """Correct Divisions sheet with detailed change tracking"""
        print("Correcting Divisions...")
        
        # Get organization names for reporting
        org_col = None
        for col in df.columns:
            if 'organization' in col.lower() or 'company' in col.lower():
                org_col = col
                break
        
        # Find relevant columns
        org_short_name_col = None
        division_name_col = None
        purpose_col = None
        first_name_col = None
        last_name_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'organization' in col_lower and ('short' in col_lower or 'short name' in col_lower):
                org_short_name_col = col
            elif 'division' in col_lower and 'name' in col_lower:
                division_name_col = col
            elif 'purpose' in col_lower:
                purpose_col = col
            elif 'first' in col_lower and 'name' in col_lower:
                first_name_col = col
            elif 'last' in col_lower and 'name' in col_lower:
                last_name_col = col
        
        # Valid purpose values
        valid_purposes = [
            'PPS-STG', 'PPS-EX-PR', 'PPS-SPO-CSPOS', 'PPS-IM-PR', 'PPS-ADMIN',
            'PPS-STPOVR', 'PPS-IM-EX', 'PPS-YO-CC', 'PPS-YO-ECS', 'PPS-HRM', 'PPS-FMG'
        ]
        
        # Process each column based on processing options
        for col in df.columns:
            col_lower = col.lower()
            
            # Organization Short Name corrections
            if org_short_name_col and col == org_short_name_col:
                # Check if corrections are enabled - default to False
                apply_corrections = False
                apply_dummy_data = False
                
                if processing_options and 'divisions' in processing_options:
                    div_options = processing_options['divisions']
                    if 'Organization Short Name' in div_options:
                        apply_corrections = div_options['Organization Short Name']['correct']
                        apply_dummy_data = div_options['Organization Short Name']['dummy_data']
                
                if apply_corrections or apply_dummy_data:
                    for idx in range(len(df)):
                        original_value = df.loc[idx, col]
                        if pd.isna(original_value) or str(original_value).strip() == '':
                            new_value = 'EmptyDummy'
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            self.log_detailed_change('Divisions', 'Organization Short Name Correction', f'Row {idx + 4}', 
                                                   col, original_value, new_value, org_name)
                            df.loc[idx, col] = new_value
            
            # Division Name corrections
            elif division_name_col and col == division_name_col:
                # Check if corrections are enabled - default to False
                apply_corrections = False
                apply_dummy_data = False
                
                if processing_options and 'divisions' in processing_options:
                    div_options = processing_options['divisions']
                    if 'Division Name' in div_options:
                        apply_corrections = div_options['Division Name']['correct']
                        apply_dummy_data = div_options['Division Name']['dummy_data']
                
                if apply_corrections or apply_dummy_data:
                    for idx in range(len(df)):
                        original_value = df.loc[idx, col]
                        if pd.isna(original_value) or str(original_value).strip() == '':
                            new_value = 'Admin'
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            self.log_detailed_change('Divisions', 'Division Name Correction', f'Row {idx + 4}', 
                                                   col, original_value, new_value, org_name)
                            df.loc[idx, col] = new_value
            
            # Purpose corrections  
            elif purpose_col and col == purpose_col:
                # Check if corrections are enabled - default to False
                apply_corrections = False
                apply_dummy_data = False
                
                if processing_options and 'divisions' in processing_options:
                    div_options = processing_options['divisions']
                    if 'Purpose' in div_options:
                        apply_corrections = div_options['Purpose']['correct']
                        apply_dummy_data = div_options['Purpose']['dummy_data']
                
                if apply_corrections or apply_dummy_data:
                    for idx in range(len(df)):
                        original_value = df.loc[idx, col]
                        if pd.isna(original_value) or str(original_value).strip() == '':
                            new_value = 'PPS-STG'
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            self.log_detailed_change('Divisions', 'Purpose Correction', f'Row {idx + 4}', 
                                                   col, original_value, new_value, org_name)
                            df.loc[idx, col] = new_value
                        elif pd.notna(original_value):
                            # Validate and correct invalid purpose values
                            purpose_str = str(original_value).strip()
                            if ',' in purpose_str:
                                # Multiple purposes - validate each one
                                purposes = [p.strip() for p in purpose_str.split(',')]
                                valid_purposes_found = [p for p in purposes if p in valid_purposes]
                                if valid_purposes_found:
                                    new_value = ', '.join(valid_purposes_found)
                                else:
                                    new_value = 'PPS-STG'
                                
                                if new_value != purpose_str:
                                    org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                    self.log_detailed_change('Divisions', 'Purpose Validation Correction', f'Row {idx + 4}', 
                                                           col, purpose_str, new_value, org_name)
                                    df.loc[idx, col] = new_value
                            else:
                                # Single purpose
                                if purpose_str not in valid_purposes:
                                    new_value = 'PPS-STG'
                                    org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                    self.log_detailed_change('Divisions', 'Purpose Validation Correction', f'Row {idx + 4}', 
                                                           col, purpose_str, new_value, org_name)
                                    df.loc[idx, col] = new_value
            
            # Principle Contact's First Name corrections
            elif first_name_col and col == first_name_col:
                # Check if corrections are enabled - default to False
                apply_corrections = False
                apply_dummy_data = False
                
                if processing_options and 'divisions' in processing_options:
                    div_options = processing_options['divisions']
                    if 'Principle Contact First Name' in div_options:
                        apply_corrections = div_options['Principle Contact First Name']['correct']
                        apply_dummy_data = div_options['Principle Contact First Name']['dummy_data']
                
                if apply_corrections or apply_dummy_data:
                    for idx in range(len(df)):
                        original_value = df.loc[idx, col]
                        if pd.isna(original_value) or str(original_value).strip() == '':
                            # Fill empty values with 'Admin'
                            new_value = 'Admin'
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            self.log_detailed_change('Divisions', 'Principle Contact First Name Correction', f'Row {idx + 4}', 
                                                   col, original_value, new_value, org_name)
                            df.loc[idx, col] = new_value
                        else:
                            # Apply concatenation rule: First Name + Organization Short Name
                            if org_short_name_col and not pd.isna(df.loc[idx, org_short_name_col]):
                                org_short_name = str(df.loc[idx, org_short_name_col]).strip()
                                if org_short_name and org_short_name != '':
                                    new_value = f"{str(original_value).strip()} {org_short_name}"
                                    if new_value != str(original_value).strip():
                                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                        self.log_detailed_change('Divisions', 'Principle Contact First Name Concatenation', f'Row {idx + 4}', 
                                                               col, original_value, new_value, org_name)
                                        df.loc[idx, col] = new_value
            
            # Principle Contact's Last Name corrections
            elif last_name_col and col == last_name_col:
                # Check if corrections are enabled - default to False
                apply_corrections = False
                apply_dummy_data = False
                
                if processing_options and 'divisions' in processing_options:
                    div_options = processing_options['divisions']
                    if 'Principle Contact Last Name' in div_options:
                        apply_corrections = div_options['Principle Contact Last Name']['correct']
                        apply_dummy_data = div_options['Principle Contact Last Name']['dummy_data']
                
                if apply_corrections or apply_dummy_data:
                    for idx in range(len(df)):
                        original_value = df.loc[idx, col]
                        if pd.isna(original_value) or str(original_value).strip() == '':
                            # Fill empty values with 'Admin'
                            new_value = 'Admin'
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            self.log_detailed_change('Divisions', 'Principle Contact Last Name Correction', f'Row {idx + 4}', 
                                                   col, original_value, new_value, org_name)
                            df.loc[idx, col] = new_value
                        else:
                            # Apply concatenation rule: Last Name + Organization Short Name
                            if org_short_name_col and not pd.isna(df.loc[idx, org_short_name_col]):
                                org_short_name = str(df.loc[idx, org_short_name_col]).strip()
                                if org_short_name and org_short_name != '':
                                    new_value = f"{str(original_value).strip()} {org_short_name}"
                                    if new_value != str(original_value).strip():
                                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                                        self.log_detailed_change('Divisions', 'Principle Contact Last Name Concatenation', f'Row {idx + 4}', 
                                                               col, original_value, new_value, org_name)
                                        df.loc[idx, col] = new_value
        
        return df
    
    def correct_human_resources(self, df, workbook=None, sheet_name=None, processing_options=None):
        """Correct Human Resources sheet while preserving red cell formatting"""
        print("Correcting Human Resources...")
        
        # Store red cell information for formatting preservation
        red_cell_info = {}
        
        # If workbook is provided, analyze red cells in the original sheet
        if workbook and sheet_name:
            original_sheet = workbook[sheet_name]
            
            # Find NIC and Email columns in the original sheet
            header_row = 3  # Headers are in row 3
            nic_col_idx = None
            email_col_idx = None
            
            for col_idx in range(1, original_sheet.max_column + 1):
                header_cell = original_sheet.cell(row=header_row, column=col_idx)
                if header_cell.value:
                    header_value = str(header_cell.value).lower()
                    if 'nic' in header_value:
                        nic_col_idx = col_idx
                    elif 'email' in header_value:
                        email_col_idx = col_idx
            
            # Analyze red cells and store their positions
            if nic_col_idx or email_col_idx:
                data_start_row = 4  # Data starts from row 4
                
                for row_idx in range(data_start_row, original_sheet.max_row + 1):
                    df_row_idx = row_idx - data_start_row  # Convert to DataFrame index
                    
                    if df_row_idx >= len(df):
                        break
                    
                    nic_is_red = False
                    email_is_red = False
                    
                    if nic_col_idx:
                        nic_cell = original_sheet.cell(row=row_idx, column=nic_col_idx)
                        nic_is_red = self.is_red_cell(nic_cell)
                    
                    if email_col_idx:
                        email_cell = original_sheet.cell(row=row_idx, column=email_col_idx)
                        email_is_red = self.is_red_cell(email_cell)
                    
                    # Store red cell information
                    red_cell_info[df_row_idx] = {
                        'nic_red': nic_is_red,
                        'email_red': email_is_red,
                        'excel_row': row_idx
                    }
        
        # Generate organization short names for dummy data
        org_short_names = []
        org_col = None
        for col in df.columns:
            if 'organization' in col.lower() or 'company' in col.lower():
                org_col = col
                break
        
        if org_col:
            org_short_names = [self.generate_org_short_name(org) for org in df[org_col]]
        else:
            org_short_names = ['ORG'] * len(df)
        
        # Process each row with red cell awareness
        for idx in range(len(df)):
            if idx in red_cell_info:
                red_info = red_cell_info[idx]
                nic_red = red_info['nic_red']
                email_red = red_info['email_red']
                excel_row = red_info['excel_row']
                
                org_name = org_short_names[idx] if idx < len(org_short_names) else 'Unknown'
                
                # Find NIC and Email columns in DataFrame
                nic_col = None
                email_col = None
                for col in df.columns:
                    if 'nic' in col.lower():
                        nic_col = col
                    elif 'email' in col.lower():
                        email_col = col
                
                # Apply red cell logic
                if nic_red and email_red:
                    # Both are red - set Activity to Update but don't change NIC/Email values
                    for col in df.columns:
                        if 'activity' in col.lower():
                            df.loc[idx, col] = 'Update'
                    
                    self.log_hr_red_cell_change('Both Red - No Change', excel_row, org_name)
                
                elif nic_red and not email_red:
                    # Only NIC is red - add prefix to NIC
                    if nic_col and pd.notna(df.loc[idx, nic_col]):
                        original_nic = str(df.loc[idx, nic_col])
                        new_nic = f"BULK{original_nic}"
                        df.loc[idx, nic_col] = new_nic
                        self.log_hr_red_cell_change('NIC Red - Prefix', excel_row, org_name, original_nic, new_nic)
                
                elif not nic_red and email_red:
                    # Only Email is red - add prefix to Email
                    if email_col and pd.notna(df.loc[idx, email_col]):
                        original_email = str(df.loc[idx, email_col])
                        new_email = f"BULK{original_email}"
                        df.loc[idx, email_col] = new_email
                        self.log_hr_red_cell_change('Email Red - Prefix', excel_row, org_name, original_email, new_email)
        
        # Handle duplicates and empty fields for NIC and Email (keeping current logic)
        if 'NIC' in df.columns or any('nic' in col.lower() for col in df.columns):
            nic_col = next((col for col in df.columns if 'nic' in col.lower()), None)
            if nic_col:
                df = self.handle_duplicates_and_empty(df, nic_col, "DUMMY", org_short_names, is_email=False)
        
        if 'Email' in df.columns or any('email' in col.lower() for col in df.columns):
            email_col = next((col for col in df.columns if 'email' in col.lower()), None)
            if email_col:
                df = self.handle_duplicates_and_empty(df, email_col, "DUMMY", org_short_names, is_email=True)
        
        # NEW VALIDATION LOGIC - Apply only if processing options allow
        if processing_options and 'human_resources' in processing_options:
            hr_options = processing_options['human_resources']
            
            # 1. First Name validation and correction
            if 'First Name' in hr_options and hr_options['First Name']['correct']:
                for col in df.columns:
                    if 'first name' in col.lower() and 'last' not in col.lower():
                        for idx in range(len(df)):
                            original_value = df.loc[idx, col]
                            if pd.isna(original_value) or str(original_value).strip() == '':
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                new_value = f"First Name {org_name}"
                                self.log_detailed_change('Human Resources', 'First Name Fill', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                df.loc[idx, col] = new_value
            
            # 2. Last Name validation and correction
            if 'Last Name' in hr_options and hr_options['Last Name']['correct']:
                for col in df.columns:
                    if 'last name' in col.lower() and 'first' not in col.lower():
                        for idx in range(len(df)):
                            original_value = df.loc[idx, col]
                            if pd.isna(original_value) or str(original_value).strip() == '':
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                new_value = f"Last Name {org_name}"
                                self.log_detailed_change('Human Resources', 'Last Name Fill', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                df.loc[idx, col] = new_value
            
            # 3. Role validation and correction
            if 'Role' in hr_options and hr_options['Role']['correct']:
                for col in df.columns:
                    if 'role' in col.lower():
                        for idx in range(len(df)):
                            original_value = df.loc[idx, col]
                            if pd.isna(original_value) or str(original_value).strip() == '':
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                new_value = "Driver"
                                self.log_detailed_change('Human Resources', 'Role Fill', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                df.loc[idx, col] = new_value
            
            # 4. Division validation and correction (always corrected to Admin)
            if 'Division' in hr_options and hr_options['Division']['correct']:
                for col in df.columns:
                    if 'division' in col.lower():
                        for idx in range(len(df)):
                            original_value = df.loc[idx, col]
                            if pd.notna(original_value) and str(original_value).strip() != 'Admin':
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                self.log_detailed_change('Human Resources', 'Division Correction', f'Row {idx + 4}', 
                                                       col, original_value, 'Admin', org_name)
                        df[col] = 'Admin'
            
            # 5. Designation validation and correction
            if 'Designation' in hr_options and hr_options['Designation']['correct']:
                for col in df.columns:
                    if 'designation' in col.lower():
                        for idx in range(len(df)):
                            original_value = df.loc[idx, col]
                            if pd.isna(original_value) or str(original_value).strip() == '':
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                new_value = "Manager"
                                self.log_detailed_change('Human Resources', 'Designation Fill', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                df.loc[idx, col] = new_value
            
            # 6. Operations validation and correction
            if 'Operations' in hr_options and hr_options['Operations']['correct']:
                for col in df.columns:
                    if 'operation' in col.lower():
                        for idx in range(len(df)):
                            original_value = df.loc[idx, col]
                            if pd.isna(original_value) or str(original_value).strip() == '':
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                new_value = "Default"
                                self.log_detailed_change('Human Resources', 'Operations Fill', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                df.loc[idx, col] = new_value
            
            # 7. Create a User Account validation and correction
            if 'Create a User Account' in hr_options and hr_options['Create a User Account']['correct']:
                for col in df.columns:
                    if 'create' in col.lower() and 'user' in col.lower() and 'account' in col.lower():
                        for idx in range(len(df)):
                            original_value = df.loc[idx, col]
                            if pd.isna(original_value) or str(original_value).strip() == '':
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                new_value = "TRUE"
                                self.log_detailed_change('Human Resources', 'Create User Account Fill', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                # Ensure the value is stored as a string to prevent boolean conversion
                                df.loc[idx, col] = str(new_value)
        
        # ALWAYS convert boolean values to text in Create a User Account column (regardless of processing options)
        for col in df.columns:
            if 'create' in col.lower() and 'user' in col.lower() and 'account' in col.lower():
                print(f"🔧 Processing Create a User Account column: {col}")
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value):
                        # Convert any existing boolean values back to text format
                        value_str = str(original_value).strip()
                        print(f"  Row {idx + 4}: Original value '{original_value}' (type: {type(original_value)}) -> String: '{value_str}'")
                        
                        # Handle both string and numeric boolean representations (including floats like 1.0, 0.0)
                        if (value_str in ['1', 'True', 'true', 'TRUE'] or 
                            original_value == 1 or original_value == 1.0 or 
                            value_str == '1.0'):
                            new_value = "TRUE"
                            # Always convert numeric values to text, even if they look like the target
                            if (original_value in [1, 1.0] or value_str in ['1', '1.0']):
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                self.log_detailed_change('Human Resources', 'Create User Account Text Conversion', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                df.loc[idx, col] = new_value
                                print(f"    Converted '{original_value}' → '{new_value}'")
                            else:
                                print(f"    Value '{original_value}' is already TRUE")
                        elif (value_str in ['0', 'False', 'false', 'FALSE'] or 
                              original_value == 0 or original_value == 0.0 or 
                              value_str == '0.0'):
                            new_value = "FALSE"
                            # Always convert numeric values to text, even if they look like the target
                            if (original_value in [0, 0.0] or value_str in ['0', '0.0']):
                                org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                                self.log_detailed_change('Human Resources', 'Create User Account Text Conversion', f'Row {idx + 4}', 
                                                       col, original_value, new_value, org_name)
                                df.loc[idx, col] = new_value
                                print(f"    Converted '{original_value}' → '{new_value}'")
                            else:
                                print(f"    Value '{original_value}' is already FALSE")
                        else:
                            print(f"    Value '{value_str}' is already in correct format")
                
                # Force the entire column to be string type to prevent pandas from converting back to boolean
                print(f"🔧 Converting column '{col}' to string type to prevent boolean conversion")
                df[col] = df[col].astype(str)
                
                # Double-check that all values are now strings
                for idx in range(len(df)):
                    final_value = df.loc[idx, col]
                    print(f"  Row {idx + 4}: Final value '{final_value}' (type: {type(final_value)})")
        
        # Gender standardization - NEW: Fill empty Gender fields with "Male"
        for col in df.columns:
            if 'gender' in col.lower():
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.isna(original_value) or str(original_value).strip() == '':
                        # Fill empty Gender fields with "Male"
                        org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                        new_value = 'Male'
                        self.log_detailed_change('Human Resources', 'Gender Fill', f'Row {idx + 4}', 
                                               col, original_value, new_value, org_name)
                        df.loc[idx, col] = new_value
                    elif pd.notna(original_value):
                        # Standardize existing values
                        if str(original_value).lower().startswith('m'):
                            new_value = 'Male'
                        elif str(original_value).lower().startswith('f'):
                            new_value = 'Female'
                        else:
                            new_value = 'Male'  # Default
                        
                        if str(original_value).strip() != new_value:
                            org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                            self.log_detailed_change('Human Resources', 'Gender Standardization', f'Row {idx + 4}', 
                                                   col, original_value, new_value, org_name)
                        df.loc[idx, col] = new_value
        
        # Fill empty principle contact names
        for col in df.columns:
            if 'principle contact' in col.lower() and 'first name' in col.lower():
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.isna(original_value) or str(original_value).strip() == '':
                        org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                        new_value = "Principle Contact's First Name"
                        self.log_detailed_change('Human Resources', 'Principle Contact First Name Fill', f'Row {idx + 4}', 
                                               col, original_value, new_value, org_name)
                        df.loc[idx, col] = new_value
                        self.log_principle_contact_change('First Name', org_name, 'Human Resources')
            
            elif 'principle contact' in col.lower() and 'last name' in col.lower():
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.isna(original_value) or str(original_value).strip() == '':
                        org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                        new_value = "Principle Contact's Last Name"
                        self.log_detailed_change('Human Resources', 'Principle Contact Last Name Fill', f'Row {idx + 4}', 
                                               col, original_value, new_value, org_name)
                        df.loc[idx, col] = new_value
                        self.log_principle_contact_change('Last Name', org_name, 'Human Resources')
        
        # Change status/activity to Create (unless already set to Update due to red cells)
        for col in df.columns:
            if 'status' in col.lower() or col == 'Activity':
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if original_value != 'Update' and (pd.isna(original_value) or str(original_value).strip() != 'Create'):
                        org_name = org_short_names[idx] if idx < len(org_short_names) else f'Row {idx + 4}'
                        new_value = 'Create'
                        self.log_detailed_change('Human Resources', 'Status Correction', f'Row {idx + 4}', 
                                               col, original_value, new_value, org_name)
                        df.loc[idx, col] = new_value
        
        return df
    
    def correct_vehicles(self, df, hr_data=None, processing_options=None):
        """Correct Vehicles sheet with detailed change tracking and HR NIC matching"""
        print("Correcting Vehicles...")
        
        # Get organization names for reporting
        org_col = None
        org_short_name_col = None
        managed_by_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'organization' in col_lower and 'short' not in col_lower:
                org_col = col
            elif 'organization' in col_lower and 'short' in col_lower:
                org_short_name_col = col
            elif 'managed by' in col_lower:
                managed_by_col = col
        
        # Standard corrections with detailed tracking
        for col in df.columns:
            # Division corrections
            if 'division' in col.lower():
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value) and str(original_value).strip() != 'Admin':
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Vehicles', 'Division Correction', f'Row {idx + 4}', 
                                               col, original_value, 'Admin', org_name)
                df[col] = 'Admin'
            
            # Vehicle Type corrections
            elif 'vehicle' in col.lower() and 'type' in col.lower():
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value) and str(original_value).strip() != 'TRUCK':
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Vehicles', 'Vehicle Type Correction', f'Row {idx + 4}', 
                                               col, original_value, 'TRUCK', org_name)
                df[col] = 'TRUCK'
            
            # Load Type corrections
            elif 'load' in col.lower() and 'type' in col.lower():
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value) and str(original_value).strip() != 'LOADS':
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Vehicles', 'Load Type Correction', f'Row {idx + 4}', 
                                               col, original_value, 'LOADS', org_name)
                df[col] = 'LOADS'
            
            # Categories formatting (e.g., 20Ft, 40Ft)
            elif 'categor' in col.lower():
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value):
                        formatted_value = self.format_vehicle_category(str(original_value))
                        if str(original_value).strip() != formatted_value:
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            self.log_detailed_change('Vehicles', 'Category Formatting', f'Row {idx + 4}', 
                                                   col, original_value, formatted_value, org_name)
                            df.loc[idx, col] = formatted_value
            
            # Status corrections
            elif 'status' in col.lower() or col == 'Activity':
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value) and str(original_value).strip() != 'Create':
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Vehicles', 'Status Correction', f'Row {idx + 4}', 
                                               col, original_value, 'Create', org_name)
                df[col] = 'Create'
        
        # Fill empty "Managed By" fields with NICs from HR data
        if managed_by_col and org_short_name_col and hr_data:
            print("Analyzing empty 'Managed By' fields for NIC matching...")
            
            for idx in range(len(df)):
                # Skip END rows
                if self.is_end_row(df, idx):
                    continue
                
                original_managed_by = df.loc[idx, managed_by_col]
                
                # Check if Managed By is empty
                if pd.isna(original_managed_by) or str(original_managed_by).strip() == '':
                    target_org_short_name = df.loc[idx, org_short_name_col]
                    
                    if pd.notna(target_org_short_name) and str(target_org_short_name).strip():
                        # Find matching NIC from HR data
                        match_result = self.find_matching_nic_from_hr(str(target_org_short_name).strip(), hr_data)
                        
                        if match_result:
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            
                            # Update the Managed By field with the matched NIC
                            df.loc[idx, managed_by_col] = match_result['nic']
                            
                            # Log the detailed change
                            self.log_detailed_change('Vehicles', 'Managed By NIC Fill', f'Row {idx + 4}', 
                                                   managed_by_col, 'Empty', match_result['nic'], org_name)
                            
                            print(f"Filled Managed By for row {idx + 4}: {match_result['nic']} ({match_result['name']}) - {match_result['reason']}")
        
        return df
    
    def handle_location_reference_id(self, df, column_name, org_short_names):
        """Handle empty and duplicate Location Reference IDs"""
        # Convert column to object type to handle mixed data types
        df[column_name] = df[column_name].astype('object')
        
        seen_values = {}
        lrid_counter = 1
        
        for idx, value in enumerate(df[column_name]):
            if pd.isna(value) or value == '':
                # Fill empty values: Organization Short Name + LRID + n
                df.loc[idx, column_name] = f"{org_short_names[idx]}LRID{lrid_counter}"
                lrid_counter += 1
            else:
                value_str = str(value).strip()
                if value_str in seen_values:
                    # Handle duplicates - keep first occurrence, change others
                    count = seen_values[value_str] + 1
                    seen_values[value_str] = count
                    df.loc[idx, column_name] = f"{value_str}DUPLICATELRID{count}"
                else:
                    # First occurrence - keep original value
                    seen_values[value_str] = 1
        
        return df
    
    def handle_location_name(self, df, column_name, org_short_names):
        """Handle empty Location Names"""
        # Convert column to object type to handle mixed data types
        df[column_name] = df[column_name].astype('object')
        
        for idx, value in enumerate(df[column_name]):
            if pd.isna(value) or value == '':
                # Fill empty values: Organization Short Name + LOC
                df.loc[idx, column_name] = f"{org_short_names[idx]}LOC"
        
        return df
    
    def correct_locations(self, df, hr_data=None, processing_options=None):
        """Correct Locations sheet with detailed change tracking and HR NIC matching"""
        print("Correcting Locations...")
        
        # Get organization names for reporting
        org_col = None
        org_short_name_col = None
        principle_contact_nic_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'organization' in col_lower and 'short' not in col_lower:
                org_col = col
            elif 'organization' in col_lower and 'short' in col_lower:
                org_short_name_col = col
            elif 'principle' in col_lower and 'contact' in col_lower and 'nic' in col_lower:
                principle_contact_nic_col = col
        
        # Handle Location Reference IDs
        location_ref_id_col = None
        location_name_col = None
        
        for col in df.columns:
            if 'location reference id' in col.lower():
                location_ref_id_col = col
            elif 'location name' in col.lower():
                location_name_col = col
        
        location_counter = 1
        duplicate_counter = 1
        
        if location_ref_id_col:
            for idx in range(len(df)):
                original_value = df.loc[idx, location_ref_id_col]
                org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                
                if pd.isna(original_value) or str(original_value).strip() == '':
                    # Get org short name for LRID
                    org_short_name = df.loc[idx, org_short_name_col] if org_short_name_col else 'ORG'
                    if pd.isna(org_short_name):
                        org_short_name = 'ORG'
                    
                    new_lrid = f"{org_short_name}LRID{location_counter}"
                    df.loc[idx, location_ref_id_col] = new_lrid
                    
                    self.log_detailed_change('Locations', 'Location Reference ID Fill', f'Row {idx + 4}', 
                                           location_ref_id_col, original_value, new_lrid, org_name)
                    location_counter += 1
                    
                elif 'DUPLICATE' in str(original_value):
                    new_lrid = f"DUPLICATELRID{duplicate_counter}"
                    self.log_detailed_change('Locations', 'Duplicate LRID Correction', f'Row {idx + 4}', 
                                           location_ref_id_col, original_value, new_lrid, org_name)
                    df.loc[idx, location_ref_id_col] = new_lrid
                    duplicate_counter += 1
        
        # Handle Location Names
        if location_name_col:
            for idx in range(len(df)):
                original_value = df.loc[idx, location_name_col]
                org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                
                if pd.isna(original_value) or str(original_value).strip() == '':
                    # Get org short name for location name
                    org_short_name = df.loc[idx, org_short_name_col] if org_short_name_col else 'ORG'
                    if pd.isna(org_short_name):
                        org_short_name = 'ORG'
                    
                    new_location_name = f"{org_short_name}LOC"
                    df.loc[idx, location_name_col] = new_location_name
                    
                    self.log_detailed_change('Locations', 'Location Name Fill', f'Row {idx + 4}', 
                                           location_name_col, original_value, new_location_name, org_name)
        
        # Status corrections
        for col in df.columns:
            if 'status' in col.lower() or col == 'Activity':
                for idx in range(len(df)):
                    original_value = df.loc[idx, col]
                    if pd.notna(original_value) and str(original_value).strip() != 'Create':
                        org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                        self.log_detailed_change('Locations', 'Status Correction', f'Row {idx + 4}', 
                                               col, original_value, 'Create', org_name)
                df[col] = 'Create'
        
        # Fill empty "Principle Contact NIC" fields with NICs from HR data
        if principle_contact_nic_col and org_short_name_col and hr_data:
            print("Analyzing empty 'Principle Contact NIC' fields for NIC matching...")
            
            for idx in range(len(df)):
                # Skip END rows
                if self.is_end_row(df, idx):
                    continue
                
                original_principle_contact_nic = df.loc[idx, principle_contact_nic_col]
                
                # Check if Principle Contact NIC is empty
                if pd.isna(original_principle_contact_nic) or str(original_principle_contact_nic).strip() == '':
                    target_org_short_name = df.loc[idx, org_short_name_col]
                    
                    if pd.notna(target_org_short_name) and str(target_org_short_name).strip():
                        # Find matching NIC from HR data
                        match_result = self.find_matching_nic_from_hr(str(target_org_short_name).strip(), hr_data)
                        
                        if match_result:
                            org_name = df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else f'Row {idx + 4}'
                            
                            # Update the Principle Contact NIC field with the matched NIC
                            df.loc[idx, principle_contact_nic_col] = match_result['nic']
                            
                            # Log the detailed change
                            self.log_detailed_change('Locations', 'Principle Contact NIC Fill', f'Row {idx + 4}', 
                                                   principle_contact_nic_col, 'Empty', match_result['nic'], org_name)
                            
                            print(f"Filled Principle Contact NIC for row {idx + 4}: {match_result['nic']} ({match_result['name']}) - {match_result['reason']}")
        
        return df
    
    def add_end_marker(self, df):
        """Add END marker after last data column if not present"""
        # Find the last column with data
        last_col_with_data = None
        for col in df.columns:
            if df[col].notna().any():
                last_col_with_data = col
        
        # Check if END already exists
        end_exists = any('END' in str(col).upper() for col in df.columns)
        
        if not end_exists and last_col_with_data:
            # Add END column after the last data column
            last_col_index = df.columns.get_loc(last_col_with_data)
            df.insert(last_col_index + 1, 'END', '')
        
        return df
    
    def correct_excel_file(self, input_file_path, output_file_path, processing_options=None):
        """Main method to correct the Excel file while preserving formatting"""
        print(f"Starting correction of: {input_file_path}")
        
        # Reset change tracking for this file
        self.reset_change_tracking()
        
        try:
            # Load the original workbook with formatting preserved
            workbook = load_workbook(input_file_path, data_only=False)
            
            # Track HR data for NIC matching
            hr_data = None
            
            # Store processed sheets for controlled order processing
            sheets_data = {}
            
            # First pass: Read all sheets and store data
            for sheet_index, sheet_name in enumerate(workbook.sheetnames, 1):
                sheet = workbook[sheet_name]
                df = pd.read_excel(input_file_path, sheet_name=sheet_name, header=2)
                sheets_data[sheet_name] = {
                    'sheet': sheet,
                    'df': df,
                    'index': sheet_index
                }
            
            # Process sheets in specific order: Organization → Divisions → Human Resources → Vehicles → Locations
            processing_order = ['organization', 'divisions', 'human', 'vehicles', 'locations']
            
            for sheet_type in processing_order:
                for sheet_name, sheet_data in sheets_data.items():
                    sheet = sheet_data['sheet']
                    df = sheet_data['df']
                    sheet_index = sheet_data['index']
                    
                    # Match sheet type
                    name_lower = sheet_name.lower()
                    should_process = False
                    
                    if sheet_type == 'organization' and ('organization' in name_lower or 'org' in name_lower):
                        should_process = True
                    elif sheet_type == 'divisions' and 'division' in name_lower:
                        should_process = True
                    elif sheet_type == 'human' and ('human' in name_lower or 'hr' in name_lower or 'resource' in name_lower):
                        should_process = True
                    elif sheet_type == 'vehicles' and 'vehicle' in name_lower:
                        should_process = True
                    elif sheet_type == 'locations' and 'location' in name_lower:
                        should_process = True
                    
                    if should_process:
                        print(f"Processing sheet: {sheet_index} - {sheet_name}")
                        
                        # Apply corrections based on sheet type and processing options
                        if sheet_type == 'organization':
                            df = self.correct_organization_details(df, processing_options)
                        elif sheet_type == 'divisions':
                            df = self.correct_divisions(df, processing_options)
                        elif sheet_type == 'human':
                            df = self.correct_human_resources(df, workbook, sheet_name, processing_options)
                            # Extract HR data for NIC matching after processing
                            hr_data = self.extract_hr_data_for_matching(df)
                        elif sheet_type == 'vehicles':
                            df = self.correct_vehicles(df, hr_data, processing_options)
                        elif sheet_type == 'locations':
                            df = self.correct_locations(df, hr_data, processing_options)
                        
                        # Add END marker if needed
                        df = self.add_end_marker(df)
                        
                        # Update the sheet with corrected data while preserving formatting
                        self.update_sheet_with_preserved_formatting(sheet, df)
            
            # Process any remaining sheets that don't match our main types
            for sheet_name, sheet_data in sheets_data.items():
                name_lower = sheet_name.lower()
                if not any(sheet_type in name_lower for sheet_type in ['organization', 'org', 'division', 'human', 'hr', 'resource', 'vehicle', 'location']):
                    sheet = sheet_data['sheet']
                    df = sheet_data['df']
                    sheet_index = sheet_data['index']
                    
                    print(f"Processing sheet: {sheet_index} - {sheet_name}")
                    
                    # For other sheets, just change status to Create
                    for col in df.columns:
                        if 'status' in col.lower() or col == 'Activity':
                            df[col] = 'Create'
                    
                    # Add END marker if needed
                    df = self.add_end_marker(df)
                    
                    # Update the sheet with corrected data while preserving formatting
                    self.update_sheet_with_preserved_formatting(sheet, df)
            
            # Highlight errors for fields that weren't processed (if processing options provided)
            if processing_options:
                self.highlight_unprocessed_errors(workbook, sheets_data, processing_options)
                # Apply cell coloring for processed cells
                self.apply_processed_cell_coloring(workbook, sheets_data, processing_options)
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_file_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # Save the workbook with preserved formatting
            workbook.save(output_file_path)
            print(f"Corrected file saved to: {output_file_path}")
            
            # Generate and display comprehensive report
            comprehensive_report = self.generate_comprehensive_report()
            print(comprehensive_report)
            
        except Exception as e:
            print(f"Error processing file: {str(e)}")
            raise
    
    def update_sheet_with_preserved_formatting(self, sheet, df):
        """Update sheet data while preserving all original formatting"""
        # Data starts from row 4 (1-indexed) since headers are in row 3
        data_start_row = 4
        
        # Get column mapping between DataFrame and Excel sheet
        header_row = 3  # Row 3 contains the actual column headers
        col_mapping = {}
        
        # Map DataFrame columns to Excel columns
        for excel_col_idx in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=header_row, column=excel_col_idx)
            if header_cell.value and str(header_cell.value).strip():
                header_name = str(header_cell.value).strip()
                if header_name in df.columns:
                    col_mapping[header_name] = excel_col_idx
        
        # Store original red cell formatting for HR sheets (simplified approach)
        red_cells_positions = set()
        is_hr_sheet = 'human' in sheet.title.lower() or 'resource' in sheet.title.lower()
        
        if is_hr_sheet:
            # Just track which cells are red, don't store complex formatting objects
            nic_col_idx = None
            email_col_idx = None
            
            for col_name, excel_col_idx in col_mapping.items():
                if 'nic' in col_name.lower():
                    nic_col_idx = excel_col_idx
                elif 'email' in col_name.lower():
                    email_col_idx = excel_col_idx
            
            # Store positions of red cells only
            for row_idx in range(data_start_row, sheet.max_row + 1):
                if nic_col_idx:
                    nic_cell = sheet.cell(row=row_idx, column=nic_col_idx)
                    if self.is_red_cell(nic_cell):
                        red_cells_positions.add((row_idx, nic_col_idx))
                
                if email_col_idx:
                    email_cell = sheet.cell(row=row_idx, column=email_col_idx)
                    if self.is_red_cell(email_cell):
                        red_cells_positions.add((row_idx, email_col_idx))
        
        # Update existing data rows
        for row_idx, (df_idx, row_data) in enumerate(df.iterrows()):
            excel_row = data_start_row + row_idx
            
            # Make sure we don't exceed the existing data range unnecessarily
            if excel_row > sheet.max_row + 50:  # Allow some buffer for new data
                break
                
            for col_name, value in row_data.items():
                if col_name in col_mapping:
                    excel_col = col_mapping[col_name]
                    cell = sheet.cell(row=excel_row, column=excel_col)
                    
                    # Special handling for Create a User Account column to ensure text values
                    if 'create' in col_name.lower() and 'user' in col_name.lower() and 'account' in col_name.lower():
                        if pd.notna(value):
                            # Force the value to be written as text to prevent Excel from converting to boolean
                            cell_value = str(value).strip()
                            if cell_value in ['1', 'True', 'true', '1.0']:
                                cell_value = "TRUE"
                            elif cell_value in ['0', 'False', 'false', '0.0']:
                                cell_value = "FALSE"
                            cell.value = cell_value
                            print(f"🔧 Excel cell {sheet.title} Row {excel_row} Col {excel_col}: Written as text '{cell_value}'")
                        else:
                            cell.value = None
                    else:
                        # Update cell value only (let Excel preserve its own formatting)
                        if pd.notna(value):
                            cell.value = value
                        else:
                            cell.value = None
                    
                    # For red cells in HR sheets, we keep the original formatting intact
                    # since we're not recreating the workbook, the red formatting should remain
        
        # Handle new END column if it was added
        if 'END' in df.columns and 'END' not in col_mapping:
            # Find the last column with data to place END column
            last_col_with_data = 0
            for col_idx in range(1, sheet.max_column + 1):
                if sheet.cell(row=header_row, column=col_idx).value:
                    last_col_with_data = col_idx
            
            end_col = last_col_with_data + 1
            
            # Add END header
            end_header_cell = sheet.cell(row=header_row, column=end_col)
            end_header_cell.value = "END"
            
            # Copy basic formatting from adjacent header cell if available
            if last_col_with_data > 0:
                source_cell = sheet.cell(row=header_row, column=last_col_with_data)
                try:
                    # Only copy basic properties that won't cause StyleProxy issues
                    if hasattr(source_cell, 'font') and source_cell.font:
                        from openpyxl.styles import Font
                        end_header_cell.font = Font(
                            name=source_cell.font.name,
                            bold=source_cell.font.bold,
                            size=source_cell.font.size
                        )
                except:
                    pass  # If formatting copy fails, just proceed without it
        
        print(f"✅ Updated {sheet.title} with corrected data while preserving formatting")

    def log_detailed_change(self, sheet_type, change_type, row_info, field_name, original_value, new_value, organization=None):
        """Log detailed changes for visual reporting"""
        change_record = {
            'change_type': change_type,
            'row_info': row_info,
            'field_name': field_name,
            'original_value': str(original_value) if original_value is not None else 'Empty',
            'new_value': str(new_value) if new_value is not None else 'Empty',
            'organization': organization or 'Unknown'
        }
        
        sheet_key = sheet_type.lower().replace(' ', '_')
        if sheet_key in self.detailed_changes:
            self.detailed_changes[sheet_key].append(change_record)
        
        print(f"📝 {sheet_type} - {change_type}: {field_name} | '{original_value}' → '{new_value}'")

    def find_matching_nic_from_hr(self, target_org_short_name, hr_data):
        """Find matching NIC from HR data based on organization and designation priority"""
        # Find all HR records for the same organization
        matching_records = []
        
        for hr_record in hr_data:
            hr_org_short_name = hr_record.get('org_short_name', '')
            if hr_org_short_name == target_org_short_name:
                matching_records.append(hr_record)
        
        if not matching_records:
            return None
        
        # Prioritize by designation: Supervisor first, then Manager
        priority_designations = ['supervisor', 'manager']
        
        for priority_designation in priority_designations:
            for record in matching_records:
                designation = str(record.get('designation', '')).lower().strip()
                if priority_designation in designation:
                    nic = record.get('nic', '')
                    if nic and str(nic).strip() and str(nic).strip() != 'nan':
                        return {
                            'nic': str(nic).strip(),
                            'name': f"{record.get('first_name', '')} {record.get('last_name', '')}".strip(),
                            'designation': record.get('designation', ''),
                            'reason': f'Matched by org ({target_org_short_name}) - Priority: {priority_designation.title()}'
                        }
        
        # If no priority designations found, take the first available NIC
        for record in matching_records:
            nic = record.get('nic', '')
            if nic and str(nic).strip() and str(nic).strip() != 'nan':
                return {
                    'nic': str(nic).strip(),
                    'name': f"{record.get('first_name', '')} {record.get('last_name', '')}".strip(),
                    'designation': record.get('designation', ''),
                    'reason': f'Matched by org ({target_org_short_name}) - First available'
                }
        
        return None
    
    def extract_hr_data_for_matching(self, df):
        """Extract HR data for NIC matching purposes"""
        hr_data = []
        
        # Find relevant columns
        org_col = None
        nic_col = None
        first_name_col = None
        last_name_col = None
        designation_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'organization' in col_lower and 'short' in col_lower:
                org_col = col
            elif 'nic' in col_lower:
                nic_col = col
            elif 'first name' in col_lower:
                first_name_col = col
            elif 'last name' in col_lower:
                last_name_col = col
            elif 'designation' in col_lower:
                designation_col = col
        
        # Extract data
        for idx in range(len(df)):
            # Skip END rows
            if self.is_end_row(df, idx):
                continue
                
            record = {
                'org_short_name': df.loc[idx, org_col] if org_col and not pd.isna(df.loc[idx, org_col]) else '',
                'nic': df.loc[idx, nic_col] if nic_col and not pd.isna(df.loc[idx, nic_col]) else '',
                'first_name': df.loc[idx, first_name_col] if first_name_col and not pd.isna(df.loc[idx, first_name_col]) else '',
                'last_name': df.loc[idx, last_name_col] if last_name_col and not pd.isna(df.loc[idx, last_name_col]) else '',
                'designation': df.loc[idx, designation_col] if designation_col and not pd.isna(df.loc[idx, designation_col]) else ''
            }
            
            # Only add records with valid org short name
            if record['org_short_name']:
                hr_data.append(record)
        
        return hr_data

    def check_issues_only(self, input_file_path, output_directory):
        """Check for issues in the file without fixing them, highlight issues, and generate report"""
        print(f"Starting issue analysis of: {input_file_path}")
        
        # Reset issue tracking
        self.issues_found = []
        
        try:
            # Create error file directory
            error_dir = os.path.join(output_directory, "Error file")
            os.makedirs(error_dir, exist_ok=True)
            
            # Load the workbook
            workbook = load_workbook(input_file_path, data_only=False)
            
            # Generate output filename
            base_name = os.path.splitext(os.path.basename(input_file_path))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            error_file_name = f"{base_name}_issues_highlighted_{timestamp}.xlsx"
            error_file_path = os.path.join(error_dir, error_file_name)
            
            # Process each sheet and identify issues
            for sheet_index, sheet_name in enumerate(workbook.sheetnames, 1):
                print(f"Analyzing sheet: {sheet_index} - {sheet_name}")
                
                sheet = workbook[sheet_name]
                df = pd.read_excel(input_file_path, sheet_name=sheet_name, header=2)
                
                # Analyze issues based on sheet type
                if 'organization' in sheet_name.lower() or 'org' in sheet_name.lower():
                    self.analyze_organization_issues(df, sheet, 'Organization Details')
                elif 'division' in sheet_name.lower():
                    self.analyze_divisions_issues(df, sheet, 'Divisions')
                elif 'human' in sheet_name.lower() or 'hr' in sheet_name.lower() or 'resource' in sheet_name.lower():
                    self.analyze_human_resources_issues(df, sheet, 'Human Resources')
                elif 'vehicle' in sheet_name.lower():
                    self.analyze_vehicles_issues(df, sheet, 'Vehicles')
                elif 'location' in sheet_name.lower():
                    self.analyze_locations_issues(df, sheet, 'Locations')
            
            # Save the workbook with highlighted issues
            workbook.save(error_file_path)
            print(f"Issues file saved to: {error_file_path}")
            
            # Generate detailed issues report
            issues_report = self.generate_issues_report()
            
            return error_file_path, issues_report
            
        except Exception as e:
            print(f"Error during issue analysis: {str(e)}")
            raise
    
    def analyze_organization_issues(self, df, sheet, sheet_type):
        """Analyze Organization Details sheet for issues with enhanced validation"""
        data_start_row = 4  # Data starts from row 4 (1-indexed)
        
        # Find relevant columns
        org_name_col = None
        org_short_name_col = None
        operations_col = None
        status_col = None
        verticals_col = None
        country_col = None
        state_col = None
        principle_contact_first_name_col = None
        principle_contact_last_name_col = None
        address_line_col = None
        city_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'organization' in col_lower and 'name' in col_lower and 'short' not in col_lower:
                org_name_col = col
            elif 'organization' in col_lower and 'short' in col_lower:
                org_short_name_col = col
            elif 'operation' in col_lower:
                operations_col = col
            elif 'status' in col_lower:
                status_col = col
            elif 'vertical' in col_lower:
                verticals_col = col
            elif 'country' in col_lower:
                country_col = col
            elif 'state' in col_lower:
                state_col = col
            elif 'principle contact' in col_lower and 'first name' in col_lower:
                principle_contact_first_name_col = col
            elif 'principle contact' in col_lower and 'last name' in col_lower:
                principle_contact_last_name_col = col
            elif 'address line' in col_lower:
                address_line_col = col
            elif 'city' in col_lower:
                city_col = col
        
        # Track organization short names for duplicate checking
        seen_org_short_names = set()
        
        # Valid values for validation
        valid_statuses = ['NON_BOI', 'BOI']
        valid_verticals = ['VERT-CUS', 'VERT-SPO', 'VERT-YO', 'VERT-IM-EX', 'VERT-SHIPPING-LINE', 'VERT-TRN']
        valid_districts = [
            'Colombo District', 'Gampaha District', 'Kalutara District', 'Kandy District', 
            'Matale District', 'Nuwara Eliya District', 'Galle District', 'Matara District', 
            'Hambantota District', 'Jaffna District', 'Kilinochchi District', 'Mannar District', 
            'Vavuniya District', 'Mullaitivu District', 'Batticaloa District', 'Ampara District', 
            'Trincomalee District', 'Kurunegala District', 'Anuradhapura District', 
            'Polonnaruwa District', 'Badulla District', 'Monaragala District', 'Ratnapura District', 
            'Kegalle District'
        ]
        
        # Check each row for issues
        for idx in range(len(df)):
            if self.is_end_row(df, idx):
                continue
                
            excel_row = data_start_row + idx
            org_name = df.loc[idx, df.columns[0]] if not pd.isna(df.loc[idx, df.columns[0]]) else f'Row {excel_row}'
            
            # 1. Organization Name cannot be empty
            if org_name_col:
                org_name_value = df.loc[idx, org_name_col]
                if pd.isna(org_name_value) or str(org_name_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, org_name_col,
                                               "Organization Name cannot be empty",
                                               sheet_type, org_name)
            
            # 2. Organization Short Name cannot be duplicated and cannot be empty
            if org_short_name_col:
                org_short_name_value = df.loc[idx, org_short_name_col]
                if pd.isna(org_short_name_value) or str(org_short_name_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, org_short_name_col,
                                               "Organization Short Name cannot be empty",
                                               sheet_type, org_name)
                else:
                    short_name_str = str(org_short_name_value).strip()
                    if short_name_str in seen_org_short_names:
                        self.add_issue_and_highlight(sheet, excel_row, org_short_name_col,
                                                   f"Duplicate Organization Short Name: '{short_name_str}'",
                                                   sheet_type, org_name)
                    else:
                        seen_org_short_names.add(short_name_str)
            
            # 3. Operations column cannot be empty
            if operations_col:
                operations_value = df.loc[idx, operations_col]
                if pd.isna(operations_value) or str(operations_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, operations_col,
                                               "Operations column cannot be empty",
                                               sheet_type, org_name)
            
            # 4. Status only can be "NON_BOI" and "BOI"
            if status_col:
                status_value = df.loc[idx, status_col]
                if pd.notna(status_value):
                    status_str = str(status_value).strip()
                    if status_str not in valid_statuses:
                        self.add_issue_and_highlight(sheet, excel_row, status_col, 
                                                   f"Status must be one of {valid_statuses}, found: '{status_str}'",
                                                   sheet_type, org_name)
                else:
                    self.add_issue_and_highlight(sheet, excel_row, status_col,
                                               "Status cannot be empty",
                                               sheet_type, org_name)
            
            # 5. Verticals are only from specific list and cannot be empty
            if verticals_col:
                verticals_value = df.loc[idx, verticals_col]
                if pd.notna(verticals_value):
                    verticals_str = str(verticals_value).strip()
                    
                    # Handle multiple verticals separated by commas
                    if ',' in verticals_str:
                        vertical_list = [v.strip() for v in verticals_str.split(',')]
                        invalid_verticals_found = []
                        
                        for vertical in vertical_list:
                            vertical = vertical.strip()
                            if vertical not in valid_verticals:
                                # Check for case-insensitive matches
                                normalized_input = vertical.upper().replace(' ', '').replace('-', '')
                                is_valid = False
                                for valid_vertical in valid_verticals:
                                    normalized_valid = valid_vertical.upper().replace(' ', '').replace('-', '')
                                    if normalized_input == normalized_valid:
                                        is_valid = True
                                        break
                                if not is_valid:
                                    invalid_verticals_found.append(vertical)
                        
                        if invalid_verticals_found:
                            self.add_issue_and_highlight(sheet, excel_row, verticals_col,
                                                       f"Invalid verticals found: {invalid_verticals_found}. Valid options: {valid_verticals}",
                                                       sheet_type, org_name)
                    else:
                        # Single vertical value
                        is_valid = False
                        if verticals_str in valid_verticals:
                            is_valid = True
                        else:
                            # Check for case-insensitive matches
                            normalized_input = verticals_str.upper().replace(' ', '').replace('-', '')
                            for valid_vertical in valid_verticals:
                                normalized_valid = valid_vertical.upper().replace(' ', '').replace('-', '')
                                if normalized_input == normalized_valid:
                                    is_valid = True
                                    break
                        
                        if not is_valid:
                            self.add_issue_and_highlight(sheet, excel_row, verticals_col,
                                                       f"Verticals must be one of {valid_verticals}, found: '{verticals_str}'",
                                                       sheet_type, org_name)
                else:
                    self.add_issue_and_highlight(sheet, excel_row, verticals_col,
                                               "Verticals cannot be empty",
                                               sheet_type, org_name)
            
            # 6. Country must be "Sri Lanka" with proper capitalization
            if country_col:
                country_value = df.loc[idx, country_col]
                if pd.notna(country_value):
                    country_str = str(country_value).strip()
                    if country_str != 'Sri Lanka':
                        self.add_issue_and_highlight(sheet, excel_row, country_col,
                                                   f"Country must be 'Sri Lanka', found: '{country_str}'",
                                                   sheet_type, org_name)
                else:
                    self.add_issue_and_highlight(sheet, excel_row, country_col,
                                               "Country cannot be empty",
                                               sheet_type, org_name)
            
            # 7. State must be from valid district list
            if state_col:
                state_value = df.loc[idx, state_col]
                if pd.notna(state_value):
                    state_str = str(state_value).strip()
                    if state_str not in valid_districts:
                        self.add_issue_and_highlight(sheet, excel_row, state_col,
                                                   f"State must be one of the valid districts, found: '{state_str}'",
                                                   sheet_type, org_name)
                else:
                    self.add_issue_and_highlight(sheet, excel_row, state_col,
                                               "State cannot be empty",
                                               sheet_type, org_name)
            
            # 8. Principle Contact's First Name cannot be empty
            if principle_contact_first_name_col:
                first_name_value = df.loc[idx, principle_contact_first_name_col]
                if pd.isna(first_name_value) or str(first_name_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, principle_contact_first_name_col,
                                               "Principle Contact's First Name cannot be empty",
                                               sheet_type, org_name)
            
            # 9. Principle Contact's Last Name cannot be empty
            if principle_contact_last_name_col:
                last_name_value = df.loc[idx, principle_contact_last_name_col]
                if pd.isna(last_name_value) or str(last_name_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, principle_contact_last_name_col,
                                               "Principle Contact's Last Name cannot be empty",
                                               sheet_type, org_name)
            
            # 10. Address Line cannot be empty
            if address_line_col:
                address_value = df.loc[idx, address_line_col]
                if pd.isna(address_value) or str(address_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, address_line_col,
                                               "Address Line cannot be empty",
                                               sheet_type, org_name)
            
            # 11. City cannot be empty
            if city_col:
                city_value = df.loc[idx, city_col]
                if pd.isna(city_value) or str(city_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, city_col,
                                               "City cannot be empty",
                                               sheet_type, org_name)
    
    def analyze_divisions_issues(self, df, sheet, sheet_type):
        """Analyze Divisions sheet for issues"""
        data_start_row = 4
        
        # Find relevant columns
        org_short_name_col = None
        division_name_col = None
        purpose_col = None
        first_name_col = None
        last_name_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'organization' in col_lower and ('short' in col_lower or 'short name' in col_lower):
                org_short_name_col = col
            elif 'division' in col_lower and 'name' in col_lower:
                division_name_col = col
            elif 'purpose' in col_lower:
                purpose_col = col
            elif 'first' in col_lower and 'name' in col_lower:
                first_name_col = col
            elif 'last' in col_lower and 'name' in col_lower:
                last_name_col = col
        
        # Valid purpose values
        valid_purposes = [
            'PPS-STG', 'PPS-EX-PR', 'PPS-SPO-CSPOS', 'PPS-IM-PR', 'PPS-ADMIN',
            'PPS-STPOVR', 'PPS-IM-EX', 'PPS-YO-CC', 'PPS-YO-ECS', 'PPS-HRM', 'PPS-FMG'
        ]
        
        # Check each row for issues
        for idx in range(len(df)):
            if self.is_end_row(df, idx):
                continue
                
            excel_row = data_start_row + idx
            org_name = df.loc[idx, df.columns[0]] if not pd.isna(df.loc[idx, df.columns[0]]) else f'Row {excel_row}'
            
            # Check Organization Short Name - cannot be empty
            if org_short_name_col:
                org_short_value = df.loc[idx, org_short_name_col]
                if pd.isna(org_short_value) or str(org_short_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, org_short_name_col,
                                               f"Organization Short Name cannot be empty",
                                               sheet_type, org_name)
            
            # Check Division Name - cannot be empty
            if division_name_col:
                division_value = df.loc[idx, division_name_col]
                if pd.isna(division_value) or str(division_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, division_name_col,
                                               f"Division Name cannot be empty",
                                               sheet_type, org_name)
            
            # Check Purpose - cannot be empty and must be valid values
            if purpose_col:
                purpose_value = df.loc[idx, purpose_col]
                if pd.isna(purpose_value) or str(purpose_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, purpose_col,
                                               f"Purpose cannot be empty",
                                               sheet_type, org_name)
                elif pd.notna(purpose_value):
                    # Check if it's a comma-separated list or single value
                    purpose_str = str(purpose_value).strip()
                    if ',' in purpose_str:
                        # Multiple purposes - check each one
                        purposes = [p.strip() for p in purpose_str.split(',')]
                        invalid_purposes = [p for p in purposes if p not in valid_purposes]
                        if invalid_purposes:
                            self.add_issue_and_highlight(sheet, excel_row, purpose_col,
                                                       f"Invalid purpose values: {', '.join(invalid_purposes)}. Valid values: {', '.join(valid_purposes)}",
                                                       sheet_type, org_name)
                    else:
                        # Single purpose
                        if purpose_str not in valid_purposes:
                            self.add_issue_and_highlight(sheet, excel_row, purpose_col,
                                                       f"Invalid purpose value: '{purpose_str}'. Valid values: {', '.join(valid_purposes)}",
                                                       sheet_type, org_name)
            
            # Check Principle Contact's First Name - cannot be empty
            if first_name_col:
                first_name_value = df.loc[idx, first_name_col]
                if pd.isna(first_name_value) or str(first_name_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, first_name_col,
                                               f"Principle Contact's First Name cannot be empty",
                                               sheet_type, org_name)
            
            # Check Principle Contact's Last Name - cannot be empty
            if last_name_col:
                last_name_value = df.loc[idx, last_name_col]
                if pd.isna(last_name_value) or str(last_name_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, last_name_col,
                                               f"Principle Contact's Last Name cannot be empty",
                                               sheet_type, org_name)
    
    def analyze_human_resources_issues(self, df, sheet, sheet_type):
        """Analyze Human Resources sheet for issues"""
        data_start_row = 4
        
        # Find relevant columns
        gender_col = None
        division_col = None
        nic_col = None
        email_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'gender' in col_lower:
                gender_col = col
            elif 'division' in col_lower:
                division_col = col
            elif 'nic' in col_lower:
                nic_col = col
            elif 'email' in col_lower:
                email_col = col
        
        # Track for duplicates
        seen_nics = set()
        seen_emails = set()
        
        # Check each row for issues
        for idx in range(len(df)):
            if self.is_end_row(df, idx):
                continue
                
            excel_row = data_start_row + idx
            org_name = df.loc[idx, df.columns[0]] if not pd.isna(df.loc[idx, df.columns[0]]) else f'Row {excel_row}'
            
            # Check gender issues
            if gender_col:
                gender_value = df.loc[idx, gender_col]
                if pd.notna(gender_value):
                    gender_str = str(gender_value).strip().lower()
                    if gender_str not in ['male', 'female']:
                        self.add_issue_and_highlight(sheet, excel_row, gender_col,
                                                   f"Gender should be 'Male' or 'Female', found: '{gender_value}'",
                                                   sheet_type, org_name)
            
            # Check division issues
            if division_col:
                division_value = df.loc[idx, division_col]
                if pd.notna(division_value) and str(division_value).strip() != 'Admin':
                    self.add_issue_and_highlight(sheet, excel_row, division_col,
                                               f"Division should be 'Admin', found: '{division_value}'",
                                               sheet_type, org_name)
            
            # Check NIC issues
            if nic_col:
                nic_value = df.loc[idx, nic_col]
                if pd.notna(nic_value):
                    nic_str = str(nic_value).strip()
                    if nic_str in seen_nics:
                        self.add_issue_and_highlight(sheet, excel_row, nic_col,
                                                   f"Duplicate NIC found: '{nic_str}'",
                                                   sheet_type, org_name)
                    else:
                        seen_nics.add(nic_str)
                else:
                    self.add_issue_and_highlight(sheet, excel_row, nic_col,
                                               "Empty NIC field",
                                               sheet_type, org_name)
            
            # Check Email issues
            if email_col:
                email_value = df.loc[idx, email_col]
                if pd.notna(email_value):
                    email_str = str(email_value).strip()
                    if email_str in seen_emails:
                        self.add_issue_and_highlight(sheet, excel_row, email_col,
                                                   f"Duplicate Email found: '{email_str}'",
                                                   sheet_type, org_name)
                    else:
                        seen_emails.add(email_str)
                else:
                    self.add_issue_and_highlight(sheet, excel_row, email_col,
                                               "Empty Email field",
                                               sheet_type, org_name)
    
    def analyze_vehicles_issues(self, df, sheet, sheet_type):
        """Analyze Vehicles sheet for issues"""
        data_start_row = 4
        
        # Find relevant columns
        division_col = None
        vehicle_type_col = None
        load_type_col = None
        managed_by_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'division' in col_lower:
                division_col = col
            elif 'vehicle' in col_lower and 'type' in col_lower:
                vehicle_type_col = col
            elif 'load' in col_lower and 'type' in col_lower:
                load_type_col = col
            elif 'managed by' in col_lower:
                managed_by_col = col
        
        # Check each row for issues
        for idx in range(len(df)):
            if self.is_end_row(df, idx):
                continue
                
            excel_row = data_start_row + idx
            org_name = df.loc[idx, df.columns[0]] if not pd.isna(df.loc[idx, df.columns[0]]) else f'Row {excel_row}'
            
            # Check division issues
            if division_col:
                division_value = df.loc[idx, division_col]
                if pd.notna(division_value) and str(division_value).strip() != 'Admin':
                    self.add_issue_and_highlight(sheet, excel_row, division_col,
                                               f"Division should be 'Admin', found: '{division_value}'",
                                               sheet_type, org_name)
            
            # Check vehicle type issues
            if vehicle_type_col:
                vehicle_type_value = df.loc[idx, vehicle_type_col]
                if pd.notna(vehicle_type_value) and str(vehicle_type_value).strip() != 'TRUCK':
                    self.add_issue_and_highlight(sheet, excel_row, vehicle_type_col,
                                               f"Vehicle Type should be 'TRUCK', found: '{vehicle_type_value}'",
                                               sheet_type, org_name)
            
            # Check load type issues
            if load_type_col:
                load_type_value = df.loc[idx, load_type_col]
                if pd.notna(load_type_value) and str(load_type_value).strip() != 'LOADS':
                    self.add_issue_and_highlight(sheet, excel_row, load_type_col,
                                               f"Load Type should be 'LOADS', found: '{load_type_value}'",
                                               sheet_type, org_name)
            
            # Check managed by issues
            if managed_by_col:
                managed_by_value = df.loc[idx, managed_by_col]
                if pd.isna(managed_by_value) or str(managed_by_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, managed_by_col,
                                               "Empty 'Managed By' field - needs NIC from HR",
                                               sheet_type, org_name)
    
    def analyze_locations_issues(self, df, sheet, sheet_type):
        """Analyze Locations sheet for issues"""
        data_start_row = 4
        
        # Find relevant columns
        location_ref_id_col = None
        location_name_col = None
        principle_contact_nic_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'location reference id' in col_lower:
                location_ref_id_col = col
            elif 'location name' in col_lower:
                location_name_col = col
            elif 'principle' in col_lower and 'contact' in col_lower and 'nic' in col_lower:
                principle_contact_nic_col = col
        
        # Check each row for issues
        for idx in range(len(df)):
            if self.is_end_row(df, idx):
                continue
                
            excel_row = data_start_row + idx
            org_name = df.loc[idx, df.columns[0]] if not pd.isna(df.loc[idx, df.columns[0]]) else f'Row {excel_row}'
            
            # Check location reference ID issues
            if location_ref_id_col:
                location_ref_value = df.loc[idx, location_ref_id_col]
                if pd.isna(location_ref_value) or str(location_ref_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, location_ref_id_col,
                                               "Empty Location Reference ID",
                                               sheet_type, org_name)
            
            # Check location name issues
            if location_name_col:
                location_name_value = df.loc[idx, location_name_col]
                if pd.isna(location_name_value) or str(location_name_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, location_name_col,
                                               "Empty Location Name",
                                               sheet_type, org_name)
            
            # Check principle contact NIC issues
            if principle_contact_nic_col:
                principle_contact_value = df.loc[idx, principle_contact_nic_col]
                if pd.isna(principle_contact_value) or str(principle_contact_value).strip() == '':
                    self.add_issue_and_highlight(sheet, excel_row, principle_contact_nic_col,
                                               "Empty 'Principle Contact NIC' field - needs NIC from HR",
                                               sheet_type, org_name)
    
    def add_issue_and_highlight(self, sheet, excel_row, column_name, issue_description, sheet_type, org_name):
        """Add issue to tracking and highlight cell in Excel"""
        # Add to issues list
        self.issues_found.append({
            'sheet': sheet_type,
            'row': excel_row,
            'column': column_name,
            'issue': issue_description,
            'organization': org_name
        })
        
        # Find the column index for highlighting
        header_row = 3  # Headers are in row 3
        col_idx = None
        
        for col in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=header_row, column=col)
            if header_cell.value and str(header_cell.value).strip() == column_name:
                col_idx = col
                break
        
        if col_idx:
            # Highlight the cell with red background
            cell = sheet.cell(row=excel_row, column=col_idx)
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            cell.fill = red_fill
            
            # Add comment with issue description
            from openpyxl.comments import Comment
            cell.comment = Comment(f"ISSUE: {issue_description}", "System")
    
    def generate_issues_report(self):
        """Generate a detailed report of all issues found"""
        if not hasattr(self, 'issues_found'):
            self.issues_found = []
        
        total_issues = len(self.issues_found)
        
        if total_issues == 0:
            return """✅ ISSUE ANALYSIS COMPLETE - NO ISSUES FOUND!

🎉 Congratulations! Your Excel file appears to be properly formatted with no issues detected.

All sheets have been analyzed and no corrections are needed:
• Organization Details: ✅ All fields properly formatted
• Divisions: ✅ All fields properly formatted  
• Human Resources: ✅ All fields properly formatted
• Vehicles: ✅ All fields properly formatted
• Locations: ✅ All fields properly formatted

Your file is ready for bulk upload as-is!"""
        
        # Group issues by sheet
        issues_by_sheet = {}
        for issue in self.issues_found:
            sheet = issue['sheet']
            if sheet not in issues_by_sheet:
                issues_by_sheet[sheet] = []
            issues_by_sheet[sheet].append(issue)
        
        report = f"""🔍 ISSUE ANALYSIS COMPLETE - {total_issues} ISSUES FOUND

================================================================================
                           ISSUES SUMMARY
================================================================================
"""
        
        for sheet, issues in issues_by_sheet.items():
            report += f"\n📋 {sheet.upper()}: {len(issues)} issues\n"
        
        report += f"\n🔴 HIGHLIGHTED CELLS: All {total_issues} issues have been highlighted in RED in the saved error file\n"
        report += "💬 COMMENTS ADDED: Each highlighted cell contains a comment explaining the specific issue\n\n"
        
        report += "================================================================================\n"
        report += "                        DETAILED ISSUES BY SHEET\n"
        report += "================================================================================\n"
        
        for sheet, issues in issues_by_sheet.items():
            report += f"\n📊 {sheet.upper()} SHEET ({len(issues)} issues):\n"
            report += "─" * 80 + "\n"
            
            for i, issue in enumerate(issues, 1):
                report += f"  {i:2}. Row {issue['row']} | {issue['column']}\n"
                report += f"      Organization: {issue['organization']}\n"
                report += f"      Issue: {issue['issue']}\n"
                report += "      " + "─" * 60 + "\n"
        
        report += "\n================================================================================\n"
        report += "                              NEXT STEPS\n"
        report += "================================================================================\n"
        report += "1. 📁 Review the highlighted error file saved to 'Error file' folder\n"
        report += "2. 🔍 Check each RED highlighted cell and its comment for specific issues\n"
        report += "3. ✏️  Manually fix issues in your original file, OR\n"
        report += "4. 🚀 Use 'Process & Fix File' button to automatically fix all issues\n"
        report += "5. ✅ Re-run analysis to confirm all issues are resolved\n\n"
        report += "💡 TIP: The 'Process & Fix File' button will automatically resolve all these issues!\n"
        
        return report
    
    def highlight_unprocessed_errors(self, workbook, sheets_data, processing_options):
        """Highlight errors in cells that weren't processed due to deselected options"""
        print("🔴 Highlighting errors in unprocessed fields...")
        
        from openpyxl.styles import PatternFill
        from openpyxl.comments import Comment
        
        # Red fill for error highlighting
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        
        for sheet_name, sheet_data in sheets_data.items():
            sheet = workbook[sheet_name]
            df = sheet_data['df']
            
            # Determine sheet type for processing options
            sheet_type = None
            name_lower = sheet_name.lower()
            
            if 'organization' in name_lower or 'org' in name_lower:
                sheet_type = 'organization'
            elif 'division' in name_lower:
                sheet_type = 'divisions'
            elif 'human' in name_lower or 'hr' in name_lower or 'resource' in name_lower:
                sheet_type = 'human_resources'
            elif 'vehicle' in name_lower:
                sheet_type = 'vehicles'
            elif 'location' in name_lower:
                sheet_type = 'locations'
            
            if sheet_type and sheet_type in processing_options:
                # Check each option for this sheet
                for option_name, option_data in processing_options[sheet_type].items():
                    # Safety check: ensure option_data has the expected structure
                    if isinstance(option_data, dict) and 'correct' in option_data:
                        # If correction is disabled, check for errors and highlight them
                        if not option_data['correct']:
                            self.highlight_field_errors(sheet, df, option_name, red_fill)
    
    def highlight_field_errors(self, sheet, df, field_name, red_fill):
        """Highlight errors for a specific field that wasn't processed"""
        # Map field names to column names and validation rules
        field_mapping = {
            # Organization Details
            'Organization Name': {'column': 'organization', 'rule': 'non_empty'},
            'Organization Short Name': {'column': 'short', 'rule': 'non_empty_unique'},
            'Operations': {'column': 'operation', 'rule': 'non_empty'},
            'Status': {'column': 'status', 'rule': 'allowed_values', 'values': ['NON_BOI', 'BOI']},
            'Verticals': {'column': 'vertical', 'rule': 'allowed_values', 'values': ['VERT-CUS', 'VERT-SPO', 'VERT-YO', 'VERT-IM-EX', 'VERT-SHIPPING-LINE', 'VERT-TRN']},
            'Country': {'column': 'country', 'rule': 'format', 'format': 'Sri Lanka'},
            'State': {'column': 'state', 'rule': 'allowed_values', 'values': ['Colombo District', 'Gampaha District', 'Kalutara District', 'Kandy District', 'Matale District', 'Nuwara Eliya District', 'Galle District', 'Matara District', 'Hambantota District', 'Jaffna District', 'Kilinochchi District', 'Mannar District', 'Vavuniya District', 'Mullaitivu District', 'Batticaloa District', 'Ampara District', 'Trincomalee District', 'Kurunegala District', 'Anuradhapura District', 'Polonnaruwa District', 'Badulla District', 'Monaragala District', 'Ratnapura District', 'Kegalle District']},
            'Principle Contact First Name': {'column': 'first', 'rule': 'non_empty'},
            'Principle Contact Last Name': {'column': 'last', 'rule': 'non_empty'},
            'Address Line': {'column': 'address', 'rule': 'non_empty'},
            'City': {'column': 'city', 'rule': 'non_empty'},
            

            
            # Human Resources
            'First Name': {'column': 'first_name', 'rule': 'non_empty'},
            'Last Name': {'column': 'last_name', 'rule': 'non_empty'},
            'Role': {'column': 'role', 'rule': 'non_empty'},
            'Division': {'column': 'division', 'rule': 'allowed_values', 'values': ['Admin']},
            'Designation': {'column': 'designation', 'rule': 'non_empty'},
            'NIC': {'column': 'nic', 'rule': 'non_empty'},
            'Email': {'column': 'email', 'rule': 'valid_email'},
            'Gender': {'column': 'gender', 'rule': 'allowed_values', 'values': ['Male', 'Female']},
            'Operations': {'column': 'operation', 'rule': 'non_empty'},
            'Create a User Account': {'column': 'create_user_account', 'rule': 'allowed_values', 'values': ['TRUE', 'FALSE']},
            'Activity': {'column': 'activity', 'rule': 'non_empty'},
            
            # Divisions
            'Organization Short Name': {'column': 'organization', 'rule': 'non_empty'},
            'Division Name': {'column': 'division', 'rule': 'non_empty'},
            'Purpose': {'column': 'purpose', 'rule': 'allowed_values', 'values': ['PPS-STG', 'PPS-EX-PR', 'PPS-SPO-CSPOS', 'PPS-IM-PR', 'PPS-ADMIN', 'PPS-STPOVR', 'PPS-IM-EX', 'PPS-YO-CC', 'PPS-YO-ECS', 'PPS-HRM', 'PPS-FMG']},
            'Principle Contact First Name': {'column': 'first', 'rule': 'non_empty'},
            'Principle Contact Last Name': {'column': 'last', 'rule': 'non_empty'},
            
            # Vehicles
            'Vehicle Type': {'column': 'type', 'rule': 'allowed_values', 'values': ['TRUCK']},
            'Load Type': {'column': 'load', 'rule': 'allowed_values', 'values': ['LOADS']},
            
            # Locations
            'Location Reference ID': {'column': 'reference', 'rule': 'non_empty'},
            'Location Name': {'column': 'name', 'rule': 'non_empty'},
            'Status': {'column': 'status', 'rule': 'allowed_values', 'values': ['Create']}
        }
        
        if field_name not in field_mapping:
            return
        
        field_config = field_mapping[field_name]
        column_key = field_config['column']
        rule = field_config['rule']
        
        # Find the actual column in the dataframe
        target_column = None
        for col in df.columns:
            col_lower = col.lower()
            if column_key in col_lower:
                target_column = col
                break
        
        if target_column is None:
            return
        
        # Check each row for errors
        for idx in range(len(df)):
            if self.is_end_row(df, idx):
                continue
            
            value = df.loc[idx, target_column]
            has_error = False
            error_description = ""
            
            # Apply validation rules
            if rule == 'non_empty':
                if pd.isna(value) or str(value).strip() == '':
                    has_error = True
                    error_description = f"Empty {field_name} field"
            
            elif rule == 'non_empty_unique':
                if pd.isna(value) or str(value).strip() == '':
                    has_error = True
                    error_description = f"Empty {field_name} field"
                # Check for duplicates (simplified)
                elif df[target_column].value_counts()[value] > 1:
                    has_error = True
                    error_description = f"Duplicate {field_name} value"
            
            elif rule == 'allowed_values':
                if pd.isna(value) or str(value).strip() == '':
                    has_error = True
                    error_description = f"Empty {field_name} field"
                elif str(value).strip() not in field_config['values']:
                    has_error = True
                    error_description = f"Invalid {field_name} value: {value}"
            
            elif rule == 'valid_email':
                if pd.isna(value) or str(value).strip() == '':
                    has_error = True
                    error_description = f"Empty {field_name} field"
                elif not self.is_valid_email(value):
                    has_error = True
                    error_description = f"Invalid email format: {value}"
            
            elif rule == 'format':
                if pd.isna(value) or str(value).strip() == '':
                    has_error = True
                    error_description = f"Empty {field_name} field"
                elif str(value).strip() != field_config['format']:
                    has_error = True
                    error_description = f"Invalid {field_name} format: expected '{field_config['format']}', got '{value}'"
            
            # Highlight error cells
            if has_error:
                # Find the cell in the sheet (accounting for header offset)
                cell_row = idx + 4  # +4 for header offset
                cell_col = None
                
                # Find column index
                for col_idx, col_name in enumerate(df.columns, 1):
                    if col_name == target_column:
                        cell_col = col_idx
                        break
                
                if cell_col is not None:
                    try:
                        cell = sheet.cell(row=cell_row, column=cell_col)
                        cell.fill = red_fill
                        
                        # Add comment explaining the issue
                        if not cell.comment:
                            cell.comment = Comment(error_description, "System")
                        else:
                            # Append to existing comment
                            existing_comment = cell.comment.text
                            cell.comment = Comment(f"{existing_comment}\n\nUNPROCESSED ISSUE: {error_description}", "System")
                        
                        print(f"🔴 Highlighted error in {sheet.title}, Row {cell_row}, Column {cell_col}: {error_description}")
                        
                    except Exception as e:
                        print(f"Warning: Could not highlight cell for {field_name}: {e}")

    def apply_processed_cell_coloring(self, workbook, sheets_data, processing_options):
        """Apply cell coloring based on processing results: red for unfixed errors only"""
        print("🎨 Applying processed cell coloring...")
        print(f"🔍 Processing options received: {processing_options}")
        
        from openpyxl.styles import PatternFill
        
        # Color definitions - only red for unfixed errors
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Red for unfixed errors
        
        for sheet_name, sheet_data in sheets_data.items():
            sheet = workbook[sheet_name]
            df = sheet_data['df']
            
            print(f"🎯 Processing sheet: {sheet_name}")
            
            # Determine sheet type for processing options
            sheet_type = None
            name_lower = sheet_name.lower()
            
            if 'organization' in name_lower or 'org' in name_lower:
                sheet_type = 'organization'
            elif 'division' in name_lower:
                sheet_type = 'divisions'
            elif 'human' in name_lower or 'hr' in name_lower or 'resource' in name_lower:
                sheet_type = 'human_resources'
            elif 'vehicle' in name_lower:
                sheet_type = 'vehicles'
            elif 'location' in name_lower:
                sheet_type = 'locations'
            
            print(f"   Sheet type determined: {sheet_type}")
            
            if sheet_type and sheet_type in processing_options:
                print(f"   Found processing options for {sheet_type}: {processing_options[sheet_type]}")
                self.color_sheet_cells(sheet, df, sheet_type, processing_options[sheet_type], red_fill)
            else:
                print(f"   No processing options found for {sheet_type}")
                if sheet_type:
                    print(f"   Available processing options: {list(processing_options.keys()) if processing_options else 'None'}")
    
    def color_sheet_cells(self, sheet, df, sheet_type, sheet_options, red_fill):
        """Color cells in a specific sheet based on processing results - only red for unfixed errors"""
        # Data starts from row 4 (1-indexed) since headers are in row 3
        data_start_row = 4
        header_row = 3
        
        # Get column mapping
        col_mapping = {}
        for excel_col_idx in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=header_row, column=excel_col_idx)
            if header_cell.value and str(header_cell.value).strip():
                header_name = str(header_cell.value).strip()
                if header_name in df.columns:
                    col_mapping[header_name] = excel_col_idx
        
        # Process each option for this sheet
        for option_name, option_data in sheet_options.items():
            if not isinstance(option_data, dict) or 'correct' not in option_data:
                continue
            
            # Find the column for this option
            target_column = self.find_column_for_option(df, option_name, sheet_type)
            if target_column is None:
                continue
            
            # Get the Excel column index
            excel_col = None
            for col_name, col_idx in col_mapping.items():
                if col_name == target_column:
                    excel_col = col_idx
                    break
            
            if excel_col is None:
                continue
            
            # Check each row and apply coloring
            for idx in range(len(df)):
                if self.is_end_row(df, idx):
                    continue
                
                excel_row = data_start_row + idx
                cell = sheet.cell(row=excel_row, column=excel_col)
                value = df.loc[idx, target_column]
                
                # Check if this cell has issues that need red coloring
                has_issues = self.cell_has_issues(value, option_name, sheet_type)
                
                if has_issues:
                    # Cell has issues - color it red
                    cell.fill = red_fill
                    print(f"🔴 Colored error cell red in {sheet.title}, Row {excel_row}, Column {excel_col}: {option_name} (Value: '{value}')")
                else:
                    # Cell is valid - leave it uncolored (white)
                    print(f"⚪ Cell not colored (valid) in {sheet.title}, Row {excel_row}, Column {excel_col}: {option_name} (Value: '{value}')")
    
    def find_column_for_option(self, df, option_name, sheet_type):
        """Find the DataFrame column that corresponds to a processing option"""
        # Map generic option names to actual column names
        column_mapping = {
            'organization': {
                'Organization Name': ['organization name', 'org name', 'name'],
                'Organization Short Name': ['organization short name', 'org short name', 'short name'],
                'Principle Contact First Name': ['principle contact first name', 'contact first name', 'first name'],
                'Principle Contact Last Name': ['principle contact last name', 'contact last name', 'last name'],
                'Principle Contact NIC': ['principle contact nic', 'contact nic', 'nic'],
                'Principle Contact Phone': ['principle contact phone', 'contact phone', 'phone'],
                'Principle Contact Email': ['principle contact email', 'contact email', 'email'],
                'Address': ['address', 'location address'],
                'City': ['city', 'city name'],
                'Province': ['province', 'state'],
                'Country': ['country', 'country name'],
                'Postal Code': ['postal code', 'zip code', 'zip'],
                'Status': ['status', 'organization status'],
                'Activity': ['activity', 'organization activity']
            },
            'divisions': {
                'Division': ['division', 'division name'],
                'Status': ['status', 'division status'],
                'Activity': ['activity', 'division activity']
            },
            'human_resources': {
                'Role': ['role', 'job role', 'position'],
                'Designation': ['designation', 'job designation', 'title'],
                'Operations': ['operations', 'operation', 'operational area'],
                'Gender': ['gender', 'sex'],
                'Create a User Account': ['create a user account', 'user account', 'account creation'],
                'Status': ['status', 'employee status'],
                'Activity': ['activity', 'employee activity']
            },
            'vehicles': {
                'Division': ['division', 'vehicle division'],
                'Vehicle Type': ['vehicle type', 'type', 'transport type'],
                'Load Type': ['load type', 'cargo type'],
                'Status': ['status', 'vehicle status'],
                'Activity': ['activity', 'vehicle activity']
            },
            'locations': {
                'Location Reference ID': ['location reference id', 'lrid', 'reference id'],
                'Location Name': ['location name', 'name'],
                'Status': ['status', 'location status'],
                'Activity': ['activity', 'location activity']
            }
        }
        
        if sheet_type in column_mapping and option_name in column_mapping[sheet_type]:
            # Look for exact matches first
            for possible_name in column_mapping[sheet_type][option_name]:
                for col in df.columns:
                    if possible_name.lower() == col.lower():
                        return col
            
            # Look for partial matches if no exact match
            for possible_name in column_mapping[sheet_type][option_name]:
                for col in df.columns:
                    if possible_name.lower() in col.lower() or col.lower() in possible_name.lower():
                        return col
        
        # Fallback: try to find by partial name matching
        for col in df.columns:
            if option_name.lower() in col.lower() or col.lower() in option_name.lower():
                return col
        
        return None
    
    def cell_has_issues(self, value, option_name, sheet_type):
        """Check if a cell has validation issues"""
        if pd.isna(value) or str(value).strip() == '':
            return True
        
        value_str = str(value).strip()
        
        # Check for specific validation rules - only for fields that have strict requirements
        if option_name == 'Gender' and value_str not in ['Male', 'Female']:
            return True
        elif option_name == 'Create a User Account' and value_str not in ['TRUE', 'FALSE']:
            # Also check for numeric representations that should be converted
            if value_str in ['1', '0', '1.0', '0.0', 'True', 'False', 'true', 'false']:
                return True  # These will be converted, so they're not final errors
            return True
        elif option_name == 'Division' and value_str != 'Admin':
            return True
        elif option_name == 'Status' and value_str not in ['Create', 'Update']:
            return True
        elif option_name == 'Activity' and value_str not in ['Create', 'Update']:
            return True
        
        # For Role, Designation, and Operations - only flag as issue if empty
        # These fields can have any non-empty value and should not be flagged as errors
        # Role: Can be "Manager", "Driver", "Transporter", etc. - only empty is an issue
        # Designation: Can be "Manager", "Transporter", etc. - only empty is an issue  
        # Operations: Can be "Cisco", "Logicare", "Kenilworth", etc. - only empty is an issue
        
        return False

def main():
    """Main function to run the correction"""
    corrector = ExcelCorrector()
    
    # Define paths
    input_file = "givenFile/DIMO-Master File Template Madumali (2).xlsx"
    
    # Create output directory if it doesn't exist
    output_dir = "Created new one"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"{output_dir}/Corrected_File_{timestamp}.xlsx"
    
    # Run correction
    try:
        corrector.correct_excel_file(input_file, output_file)
        print("Excel file correction completed successfully!")
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main() 